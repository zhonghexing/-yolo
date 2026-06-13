"""
钢材缺陷检测系统 v3.0 - 现代化 UI
Steel Defect Detection System - Modern UI

基于 UI/UX 设计原则重构：
- 深色主题 + 工业风配色
- 卡片式布局
- 现代化按钮和表格
- 流畅动画过渡
- 响应式设计
"""

import os
import sys
import time
import json
import shutil
import threading
import webbrowser
from pathlib import Path
from datetime import datetime
from typing import List, Optional

import cv2
import numpy as np

from constants import CLASS_NAMES_CN, CLASS_COLORS_RGB, CLASS_NAMES, CLASS_COLORS_BGR
from inference import _put_cn_text, draw_detections_cn
from feedback import FeedbackManager
import db

from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QLabel, QPushButton, QTableWidget, QTableWidgetItem, QHeaderView,
    QMenuBar, QMenu, QAction, QToolBar, QStatusBar, QFileDialog,
    QSplitter, QGroupBox, QGridLayout, QProgressBar, QMessageBox,
    QComboBox, QSpinBox, QDoubleSpinBox, QDialog, QFormLayout,
    QDialogButtonBox, QTextEdit, QFrame, QSizePolicy, QStyle,
    QCheckBox, QTabWidget, QScrollArea, QGraphicsDropShadowEffect
)
from PyQt5.QtCore import (
    Qt, QTimer, QSize, QThread, pyqtSignal, QMimeData, QUrl,
    QPropertyAnimation, QEasingCurve, QPoint
)
from PyQt5.QtGui import (
    QPixmap, QImage, QIcon, QFont, QColor, QPalette,
    QPainter, QPen, QBrush, QDragEnterEvent, QDropEvent,
    QLinearGradient, QRadialGradient
)


# ============================================================
# 常量定义
# ============================================================

APP_TITLE = "Steel Defect Inspector"
APP_TITLE_CN = "钢材缺陷检测系统"
APP_VERSION = "4.0.0"
WINDOW_MIN_SIZE = (1200, 800)

# ── 极简深色配色 ──
COLOR_BG             = "#1E1E1E"   # 统一背景
COLOR_BG_ALT         = "#252526"   # 次级背景
COLOR_SURFACE        = "#2D2D30"   # 表面/卡片
COLOR_BORDER         = "#3E3E42"   # 边框

COLOR_TEXT            = "#E0E0E0"   # 主文字
COLOR_TEXT_DIM        = "#969696"   # 次文字
COLOR_TEXT_FAINT      = "#5A5A5A"   # 弱文字

COLOR_ACCENT         = "#4A9EFF"   # 唯一强调色
COLOR_OK             = "#4ADE80"   # 合格
COLOR_ERR            = "#F87171"   # 不合格
COLOR_WARN           = "#FBBF24"   # 警告


# ============================================================
# 样式表 - 极简风格
# ============================================================

MODERN_STYLE = f"""
QMainWindow, QWidget {{
    background-color: {COLOR_BG};
    color: {COLOR_TEXT};
    font-family: "Microsoft YaHei", "Segoe UI", "PingFang SC", sans-serif;
    font-size: 13px;
}}

/* ── 菜单栏 ── */
QMenuBar {{
    background: {COLOR_BG_ALT};
    color: {COLOR_TEXT_DIM};
    padding: 2px 8px;
    border-bottom: 1px solid {COLOR_BORDER};
}}
QMenuBar::item {{
    padding: 6px 12px;
}}
QMenuBar::item:selected {{
    background: {COLOR_SURFACE};
    color: {COLOR_TEXT};
}}

QMenu {{
    background: {COLOR_BG_ALT};
    color: {COLOR_TEXT};
    border: 1px solid {COLOR_BORDER};
    padding: 4px;
}}
QMenu::item {{
    padding: 8px 20px;
}}
QMenu::item:selected {{
    background: {COLOR_SURFACE};
}}
QMenu::separator {{
    height: 1px;
    background: {COLOR_BORDER};
    margin: 4px 8px;
}}

/* ── 工具栏 ── */
QToolBar {{
    background: {COLOR_BG_ALT};
    border-bottom: 1px solid {COLOR_BORDER};
    spacing: 4px;
    padding: 4px 8px;
}}

/* ── 按钮 ── */
QPushButton {{
    background: {COLOR_SURFACE};
    color: {COLOR_TEXT};
    border: 1px solid {COLOR_BORDER};
    padding: 8px 16px;
    font-size: 13px;
    min-height: 18px;
}}
QPushButton:hover {{
    background: {COLOR_BORDER};
    border-color: {COLOR_TEXT_FAINT};
}}
QPushButton:pressed {{
    background: {COLOR_ACCENT};
    color: #fff;
    border-color: {COLOR_ACCENT};
}}
QPushButton:disabled {{
    color: {COLOR_TEXT_FAINT};
    border-color: {COLOR_BORDER};
}}

QPushButton#startBtn {{
    background: {COLOR_OK};
    color: #111;
    border: none;
    font-weight: 600;
    padding: 8px 24px;
}}
QPushButton#startBtn:hover {{
    background: #5cf098;
}}

QPushButton#clearBtn {{
    background: transparent;
    border: 1px solid {COLOR_BORDER};
    color: {COLOR_TEXT_DIM};
}}
QPushButton#clearBtn:hover {{
    color: {COLOR_ERR};
    border-color: {COLOR_ERR};
}}

QPushButton#timerBtn {{
    background: {COLOR_WARN};
    color: #111;
    border: none;
    font-weight: 600;
}}

QPushButton#timerPauseBtn {{
    background: {COLOR_SURFACE};
    color: {COLOR_TEXT};
    border: 1px solid {COLOR_BORDER};
    font-weight: 600;
    padding: 2px 6px;
}}

QPushButton#timerFinishBtn {{
    background: {COLOR_ACCENT};
    color: #fff;
    border: none;
    font-weight: 600;
    padding: 2px 6px;
}}
QPushButton#timerFinishBtn:hover {{
    background: #5aaeff;
}}

QPushButton#cameraBtn, QPushButton#videoBtn {{
    background: {COLOR_SURFACE};
    color: {COLOR_TEXT};
    border: 1px solid {COLOR_BORDER};
}}

QPushButton#stopVideoBtn {{
    background: {COLOR_ERR};
    color: #fff;
    border: none;
}}

QPushButton#snapshotBtn {{
    background: {COLOR_ACCENT};
    color: #fff;
    border: none;
    font-weight: 600;
}}

/* ── 卡片 ── */
QFrame#card {{
    background: {COLOR_BG_ALT};
    border: 1px solid {COLOR_BORDER};
}}

/* ── 统计卡片 ── */
QFrame#statCard {{
    background: {COLOR_BG_ALT};
    border: 1px solid {COLOR_BORDER};
}}
QFrame#statCard[cardType="pass"] {{
    border-left: 2px solid {COLOR_OK};
}}
QFrame#statCard[cardType="fail"] {{
    border-left: 2px solid {COLOR_ERR};
}}
QFrame#statCard[cardType="total"] {{
    border-left: 2px solid {COLOR_ACCENT};
}}
QFrame#statCard[cardType="rate"] {{
    border-left: 2px solid {COLOR_WARN};
}}

/* ── 表格 ── */
QTableWidget {{
    background: {COLOR_BG_ALT};
    alternate-background: {COLOR_BG_ALT};
    border: 1px solid {COLOR_BORDER};
    gridline-color: {COLOR_BORDER};
    selection-background: {COLOR_SURFACE};
    font-size: 12px;
}}
QTableWidget::item {{
    padding: 6px 8px;
    border-bottom: 1px solid {COLOR_BORDER};
    background: {COLOR_BG_ALT};
}}
QHeaderView::section {{
    background: {COLOR_SURFACE};
    color: {COLOR_TEXT_DIM};
    padding: 6px 8px;
    border: none;
    border-bottom: 1px solid {COLOR_BORDER};
    font-size: 11px;
}}

/* ── 进度条 ── */
QProgressBar {{
    background: {COLOR_SURFACE};
    height: 6px;
    text-align: center;
}}
QProgressBar::chunk {{
    background: {COLOR_ACCENT};
}}

/* ── 状态栏 ── */
QStatusBar {{
    background: {COLOR_BG_ALT};
    color: {COLOR_TEXT_FAINT};
    border-top: 1px solid {COLOR_BORDER};
    padding: 2px 8px;
    font-size: 11px;
}}

/* ── 标签 ── */
QLabel {{
    color: {COLOR_TEXT};
}}
QLabel#timerLabel {{
    font-size: 14px;
    font-weight: 600;
    color: {COLOR_WARN};
}}

/* ── 滚动条 ── */
QScrollBar:vertical {{
    background: transparent;
    width: 6px;
}}
QScrollBar::handle:vertical {{
    background: {COLOR_BORDER};
    min-height: 30px;
}}
QScrollBar::handle:vertical:hover {{
    background: {COLOR_TEXT_FAINT};
}}
QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {{
    height: 0;
}}
QScrollBar:horizontal {{
    background: transparent;
    height: 6px;
}}
QScrollBar::handle:horizontal {{
    background: {COLOR_BORDER};
    min-width: 30px;
}}

/* ── 对话框 ── */
QDialog, QMessageBox {{
    background: {COLOR_BG_ALT};
}}

/* ── 输入框 ── */
QSpinBox, QDoubleSpinBox, QComboBox {{
    background: {COLOR_SURFACE};
    color: {COLOR_TEXT};
    border: 1px solid {COLOR_BORDER};
    padding: 4px 8px;
}}
QComboBox::drop-down {{
    border: none;
    width: 20px;
}}

/* ── 文本编辑 ── */
QTextEdit {{
    background: {COLOR_SURFACE};
    color: {COLOR_TEXT};
    border: 1px solid {COLOR_BORDER};
    padding: 8px;
}}

/* ── 复选框 ── */
QCheckBox {{
    color: {COLOR_TEXT};
    spacing: 6px;
}}
QCheckBox::indicator {{
    width: 16px;
    height: 16px;
    border: 1px solid {COLOR_BORDER};
    background: {COLOR_SURFACE};
}}
QCheckBox::indicator:checked {{
    background: {COLOR_ACCENT};
    border-color: {COLOR_ACCENT};
}}

/* ── 分割器 ── */
QSplitter::handle {{
    background: {COLOR_BORDER};
}}
QSplitter::handle:hover {{
    background: {COLOR_ACCENT};
}}
"""


# ============================================================
# 统计对话框
# ============================================================

class StatsDialog(QDialog):
    """现代化统计对话框"""
    
    def __init__(self, parent=None, elapsed=0, total_seconds=180, 
                 total=0, pass_count=0, fail_count=0, defect_types=None):
        super().__init__(parent)
        self.setWindowTitle("检测统计")
        self.setMinimumWidth(400)
        self.setModal(True)
        
        # 计算合格率
        self.rate = (pass_count / total * 100) if total > 0 else 0
        self.elapsed = elapsed
        self.total_seconds = total_seconds
        self.total = total
        self.pass_count = pass_count
        self.fail_count = fail_count
        self.defect_types = defect_types or {}
        
        self._init_ui()
    
    def _init_ui(self):
        """初始化界面"""
        layout = QVBoxLayout(self)
        layout.setContentsMargins(24, 24, 24, 24)
        layout.setSpacing(16)
        
        # 标题
        title = QLabel("检测统计报告")
        title.setStyleSheet(f"""
            font-size: 18px;
            font-weight: 700;
            color: {COLOR_TEXT};
            padding-bottom: 8px;
            border-bottom: 2px solid {COLOR_ACCENT};
        """)
        layout.addWidget(title)
        
        # 用时信息
        time_frame = QFrame()
        time_frame.setStyleSheet(f"""
            QFrame {{
                background: {COLOR_SURFACE};
                border: 1px solid {COLOR_BORDER};
                border-radius: 8px;
                padding: 12px;
            }}
        """)
        time_layout = QHBoxLayout(time_frame)
        
        time_icon = QLabel("⏱")
        time_icon.setStyleSheet("font-size: 24px;")
        time_layout.addWidget(time_icon)
        
        time_info = QVBoxLayout()
        time_label = QLabel("检测用时")
        time_label.setStyleSheet(f"color: {COLOR_TEXT_DIM}; font-size: 12px;")
        time_info.addWidget(time_label)
        
        time_value = QLabel(f"{self.elapsed} 秒 / {self.total_seconds} 秒")
        time_value.setStyleSheet(f"color: {COLOR_ACCENT}; font-size: 16px; font-weight: 600;")
        time_info.addWidget(time_value)
        
        time_layout.addLayout(time_info)
        time_layout.addStretch()
        layout.addWidget(time_frame)
        
        # 统计卡片区域
        cards_layout = QHBoxLayout()
        cards_layout.setSpacing(12)
        
        # 检测总数卡片
        total_card = self._create_stat_card("检测总数", str(self.total), COLOR_ACCENT, "📊")
        cards_layout.addWidget(total_card)
        
        # 合格数卡片
        pass_card = self._create_stat_card("合格", str(self.pass_count), COLOR_OK, "✓")
        cards_layout.addWidget(pass_card)
        
        # 不合格数卡片
        fail_card = self._create_stat_card("不合格", str(self.fail_count), COLOR_ERR, "✗")
        cards_layout.addWidget(fail_card)
        
        # 合格率卡片
        rate_card = self._create_stat_card("合格率", f"{self.rate:.1f}%", 
                                          COLOR_OK if self.rate >= 90 else COLOR_WARN if self.rate >= 70 else COLOR_ERR, 
                                          "%")
        cards_layout.addWidget(rate_card)
        
        layout.addLayout(cards_layout)
        
        # 缺陷分布（如果有）
        if self.defect_types:
            defect_frame = QFrame()
            defect_frame.setStyleSheet(f"""
                QFrame {{
                    background: {COLOR_SURFACE};
                    border: 1px solid {COLOR_BORDER};
                    border-radius: 8px;
                    padding: 16px;
                }}
            """)
            defect_layout = QVBoxLayout(defect_frame)
            defect_layout.setSpacing(12)
            
            defect_title = QLabel("缺陷分布")
            defect_title.setStyleSheet(f"color: {COLOR_TEXT}; font-size: 14px; font-weight: 600;")
            defect_layout.addWidget(defect_title)
            
            # 获取缺陷颜色
            from constants import CLASS_COLORS_RGB, CLASS_NAMES_CN
            
            for class_name, count in self.defect_types.items():
                defect_row = QHBoxLayout()
                
                # 彩色标签
                color_rgb = CLASS_COLORS_RGB.get(class_name, (74, 158, 255))
                color_hex = f"#{color_rgb[0]:02x}{color_rgb[1]:02x}{color_rgb[2]:02x}"
                
                tag = QLabel()
                tag.setFixedSize(12, 12)
                tag.setStyleSheet(f"""
                    background: {color_hex};
                    border-radius: 3px;
                    border: 1px solid {color_hex};
                """)
                defect_row.addWidget(tag)
                
                # 缺陷名称（使用中文）
                cn_name = CLASS_NAMES_CN.get(class_name, class_name)
                name_label = QLabel(cn_name)
                name_label.setStyleSheet(f"color: {COLOR_TEXT}; font-size: 13px;")
                name_label.setFixedWidth(80)
                defect_row.addWidget(name_label)
                
                # 数量
                count_label = QLabel(f"{count} 个")
                count_label.setStyleSheet(f"color: {color_hex}; font-size: 14px; font-weight: 600;")
                defect_row.addWidget(count_label)
                
                # 进度条
                max_count = max(self.defect_types.values()) if self.defect_types else 1
                progress = QProgressBar()
                progress.setFixedHeight(8)
                progress.setTextVisible(False)
                progress.setMaximum(max_count)
                progress.setValue(count)
                progress.setStyleSheet(f"""
                    QProgressBar {{
                        background: {COLOR_BG_ALT};
                        border-radius: 4px;
                    }}
                    QProgressBar::chunk {{
                        background: {color_hex};
                        border-radius: 4px;
                    }}
                """)
                defect_row.addWidget(progress)
                
                defect_layout.addLayout(defect_row)
            
            layout.addWidget(defect_frame)
        
        # 关闭按钮
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        
        close_btn = QPushButton("关闭")
        close_btn.setFixedSize(100, 36)
        close_btn.setStyleSheet(f"""
            QPushButton {{
                background: {COLOR_ACCENT};
                color: white;
                border: none;
                border-radius: 6px;
                font-weight: 600;
                font-size: 14px;
            }}
            QPushButton:hover {{
                background: #5aaeff;
            }}
            QPushButton:pressed {{
                background: #3a8eef;
            }}
        """)
        close_btn.clicked.connect(self.accept)
        btn_layout.addWidget(close_btn)
        
        layout.addLayout(btn_layout)
    
    def _create_stat_card(self, title, value, color, icon):
        """创建统计卡片"""
        card = QFrame()
        card.setStyleSheet(f"""
            QFrame {{
                background: {COLOR_SURFACE};
                border: 1px solid {COLOR_BORDER};
                border-left: 3px solid {color};
                border-radius: 8px;
                padding: 12px;
            }}
        """)
        
        layout = QVBoxLayout(card)
        layout.setSpacing(4)
        
        # 图标和标题
        header = QHBoxLayout()
        icon_label = QLabel(icon)
        icon_label.setStyleSheet(f"color: {color}; font-size: 16px;")
        header.addWidget(icon_label)
        
        title_label = QLabel(title)
        title_label.setStyleSheet(f"color: {COLOR_TEXT_DIM}; font-size: 11px;")
        header.addWidget(title_label)
        header.addStretch()
        
        layout.addLayout(header)
        
        # 数值
        value_label = QLabel(value)
        value_label.setStyleSheet(f"color: {color}; font-size: 20px; font-weight: 700;")
        layout.addWidget(value_label)
        
        return card


# ============================================================
# 主应用类
# ============================================================

class SteelDefectApp(QMainWindow):
    """钢材缺陷检测主应用窗口 - 现代化 UI"""

    def __init__(self):
        super().__init__()

        # 检测器（延迟初始化）
        self.detector = None
        self.detection_thread = None

        # 视频流状态
        self.video_thread = None
        self.is_video_streaming = False
        self.current_video_frame = None
        self.current_video_result = None

        # 当前状态
        self.current_image_path = ""
        self.current_result = None
        self.all_results = []
        self.is_detecting = False

        # 图片列表导航
        self.image_list = []           # 当前文件夹中的图片路径列表
        self.current_image_index = -1  # 当前显示的图片索引

        # 批量检测队列（攒批处理，避免 UI 卡死）
        self._batch_queue = []
        self._batch_timer = QTimer(self)
        self._batch_timer.timeout.connect(self._flush_batch_queue)

        # 设置
        self.settings = {
            "model_path": "",
            "conf_threshold": 0.25,
            "iou_threshold": 0.45,
            "img_size": 640,
            "enable_voice": True,
            "enable_log": True,
            "auto_save_defects": True,
            "auto_save_dir": str(Path(__file__).parent / "defect_records"),
            "web_port": 5000,
        }
        # 自动保存目录
        Path(self.settings["auto_save_dir"]).mkdir(parents=True, exist_ok=True)
        # Web 服务线程
        self.web_thread = None

        # 反馈管理器（视觉/语音/日志三重反馈）
        self.feedback = FeedbackManager(
            enable_voice=self.settings["enable_voice"],
            enable_log=self.settings["enable_log"],
            log_dir=str(Path(__file__).parent / "runs" / "logs"),
        )

        # 灰度模式
        self.grayscale_enabled = False

        # 计时器模式
        self.timer_mode = False
        self.timer_paused = False
        self.timer_seconds = 180
        self.timer_remaining = 180
        self.countdown_timer = QTimer(self)
        self.countdown_timer.timeout.connect(self._update_timer)

        # 初始化UI
        self._init_ui()
        self._init_menu()
        self._init_toolbar()
        self._init_statusbar()
        self._connect_signals()

        # 初始化检测器
        self._init_detector()

    def _init_ui(self):
        """初始化主界面 —— 左右分栏布局"""
        self.setWindowTitle(f"{APP_TITLE} - {APP_TITLE_CN}")
        self.setMinimumSize(*WINDOW_MIN_SIZE)

        icon_path = Path(__file__).parent / "app_icon.ico"
        if icon_path.exists():
            self.setWindowIcon(QIcon(str(icon_path)))

        central_widget = QWidget()
        central_widget.setObjectName("centralWidget")
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)
        main_layout.setContentsMargins(8, 4, 8, 8)
        main_layout.setSpacing(6)

        # ── 顶部标题栏 ──
        header = QHBoxLayout()
        header.setSpacing(6)

        title = QLabel(f"{APP_TITLE_CN}")
        title.setStyleSheet(f"color: {COLOR_TEXT_DIM}; font-size: 11px;")
        title.setFixedHeight(16)
        header.addWidget(title)

        header.addStretch()

        self.timer_label = QLabel("180s")
        self.timer_label.setObjectName("timerLabel")
        self.timer_label.setFixedHeight(16)
        self.timer_label.hide()
        header.addWidget(self.timer_label)

        # 暂停按钮（放在计时器标签旁边）
        self.timer_pause_btn = QPushButton("暂停")
        self.timer_pause_btn.setObjectName("timerPauseBtn")
        self.timer_pause_btn.setFixedSize(64, 24)
        self.timer_pause_btn.clicked.connect(self._toggle_timer_pause)
        self.timer_pause_btn.hide()
        header.addWidget(self.timer_pause_btn)

        # 完成按钮（计时模式下提前结束并统计）
        self.timer_finish_btn = QPushButton("完成")
        self.timer_finish_btn.setObjectName("timerFinishBtn")
        self.timer_finish_btn.setFixedSize(64, 24)
        self.timer_finish_btn.clicked.connect(self._finish_timer)
        self.timer_finish_btn.hide()
        header.addWidget(self.timer_finish_btn)

        main_layout.addLayout(header)

        # ── 主内容区：左右分栏 ──
        content_splitter = QSplitter(Qt.Horizontal)
        content_splitter.setHandleWidth(3)

        # 左侧：预览区（图片/视频）
        preview_panel = self._create_preview_panel()
        content_splitter.addWidget(preview_panel)

        # 右侧：控制 & 结果面板
        right_panel = self._create_right_panel()
        content_splitter.addWidget(right_panel)

        content_splitter.setStretchFactor(0, 3)  # 预览占 75%
        content_splitter.setStretchFactor(1, 1)  # 右侧面板占 25%
        content_splitter.setSizes([900, 310])
        main_layout.addWidget(content_splitter)

    def _create_preview_panel(self):
        """左侧预览面板 — 最大化显示区域"""
        panel = QFrame()
        panel.setObjectName("card")
        layout = QVBoxLayout(panel)
        layout.setContentsMargins(2, 2, 2, 2)
        layout.setSpacing(2)

        # 图片/视频显示（占满可用空间）
        self.image_display = QLabel()
        self.image_display.setAlignment(Qt.AlignCenter)
        self.image_display.setMinimumSize(400, 300)
        self.image_display.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.image_display.setStyleSheet(f"""
            background-color: {COLOR_SURFACE};
            border: 1px solid {COLOR_BORDER};
            border-radius: 8px;
            padding: 16px;
            color: {COLOR_TEXT_DIM};
            font-size: 14px;
        """)
        self.image_display.setText("拖拽图片到此处\n或通过工具栏上传")
        self.image_display.setAcceptDrops(True)
        layout.addWidget(self.image_display)

        # 底部信息栏（单行：上一张 / 信息 / 下一张）
        info_bar = QHBoxLayout()
        info_bar.setSpacing(8)

        self.prev_btn = QPushButton("◀ 上一张")
        self.prev_btn.setFixedHeight(22)
        self.prev_btn.setStyleSheet(f"""
            QPushButton {{
                background: {COLOR_SURFACE};
                color: {COLOR_TEXT_DIM};
                border: 1px solid {COLOR_BORDER};
                padding: 2px 10px;
                font-size: 11px;
            }}
            QPushButton:hover {{ color: {COLOR_TEXT}; border-color: {COLOR_TEXT_FAINT}; }}
            QPushButton:disabled {{ color: {COLOR_TEXT_FAINT}; border-color: {COLOR_BORDER}; }}
        """)
        self.prev_btn.clicked.connect(self._prev_image)
        self.prev_btn.setEnabled(False)
        info_bar.addWidget(self.prev_btn)

        self.image_info_label = QLabel("")
        self.image_info_label.setStyleSheet(f"color: {COLOR_TEXT_FAINT}; font-size: 10px;")
        info_bar.addWidget(self.image_info_label, 1)

        self.next_btn = QPushButton("下一张 ▶")
        self.next_btn.setFixedHeight(22)
        self.next_btn.setStyleSheet(f"""
            QPushButton {{
                background: {COLOR_SURFACE};
                color: {COLOR_TEXT_DIM};
                border: 1px solid {COLOR_BORDER};
                padding: 2px 10px;
                font-size: 11px;
            }}
            QPushButton:hover {{ color: {COLOR_TEXT}; border-color: {COLOR_TEXT_FAINT}; }}
            QPushButton:disabled {{ color: {COLOR_TEXT_FAINT}; border-color: {COLOR_BORDER}; }}
        """)
        self.next_btn.clicked.connect(self._next_image)
        self.next_btn.setEnabled(False)
        info_bar.addWidget(self.next_btn)

        layout.addLayout(info_bar)

        return panel

    def _create_right_panel(self):
        """右侧面板 — 控制 & 结果（可滚动）"""
        outer = QFrame()
        outer.setObjectName("card")
        outer_layout = QVBoxLayout(outer)
        outer_layout.setContentsMargins(0, 0, 0, 0)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.NoFrame)
        scroll.setStyleSheet(f"QScrollArea {{ background: transparent; border: none; }}")

        panel = QWidget()
        layout = QVBoxLayout(panel)
        layout.setContentsMargins(10, 8, 10, 8)
        layout.setSpacing(6)

        # ── 1. 检测模式 ──
        mode_frame = QFrame()
        mode_frame.setObjectName("card")
        mode_layout = QVBoxLayout(mode_frame)
        mode_layout.setContentsMargins(10, 8, 10, 8)
        mode_layout.setSpacing(6)
        mode_title = QLabel("检测模式")
        mode_title.setStyleSheet(f"color: {COLOR_TEXT_DIM}; font-size: 11px; font-weight: 600;")
        mode_layout.addWidget(mode_title)
        self.right_mode_combo = QComboBox()
        self.right_mode_combo.addItems(["图片检测", "摄像头检测", "视频检测"])
        self.right_mode_combo.currentIndexChanged.connect(self._on_right_mode_changed)
        mode_layout.addWidget(self.right_mode_combo)

        # 灰度模式复选框
        self.right_grayscale_cb = QCheckBox("灰度模式")
        self.right_grayscale_cb.setStyleSheet(f"color: {COLOR_TEXT_DIM}; font-size: 11px;")
        self.right_grayscale_cb.toggled.connect(self._on_grayscale_toggled)
        mode_layout.addWidget(self.right_grayscale_cb)
        # 图片模式操作按钮
        img_actions = QHBoxLayout()
        img_actions.setSpacing(4)
        open_btn2 = QPushButton("打开图片")
        open_btn2.clicked.connect(self._open_image)
        img_actions.addWidget(open_btn2)
        open_dir_btn2 = QPushButton("文件夹")
        open_dir_btn2.clicked.connect(self._open_directory)
        img_actions.addWidget(open_dir_btn2)
        self.start_btn2 = QPushButton("开始检测")
        self.start_btn2.setObjectName("startBtn")
        self.start_btn2.clicked.connect(self._start_detection)
        self.start_btn2.setEnabled(False)
        img_actions.addWidget(self.start_btn2)
        self.clear_btn2 = QPushButton("清除")
        self.clear_btn2.setObjectName("clearBtn")
        self.clear_btn2.clicked.connect(self._clear_results)
        img_actions.addWidget(self.clear_btn2)
        self.right_image_actions = img_actions
        mode_layout.addLayout(img_actions)
        # 摄像头模式操作按钮
        cam_actions = QHBoxLayout()
        cam_actions.setSpacing(4)
        self.right_camera_combo = QComboBox()
        self.right_camera_combo.setFixedWidth(140)
        cam_actions.addWidget(self.right_camera_combo)
        self.right_camera_btn = QPushButton("启动")
        self.right_camera_btn.setObjectName("cameraBtn")
        self.right_camera_btn.clicked.connect(self._start_camera)
        cam_actions.addWidget(self.right_camera_btn)
        self.right_phone_btn = QPushButton("📱 手机")
        self.right_phone_btn.setObjectName("cameraBtn")
        self.right_phone_btn.setStyleSheet(f"""
            QPushButton {{
                background: {COLOR_ACCENT};
                color: white;
                border: none;
                padding: 4px 8px;
                font-weight: 600;
            }}
            QPushButton:hover {{
                background: #5aaeff;
            }}
        """)
        self.right_phone_btn.clicked.connect(self._start_phone_camera)
        cam_actions.addWidget(self.right_phone_btn)
        self.right_video_btn = QPushButton("选择视频")
        self.right_video_btn.setObjectName("videoBtn")
        self.right_video_btn.clicked.connect(self._open_video_file)
        cam_actions.addWidget(self.right_video_btn)
        self.right_stop_btn = QPushButton("停止")
        self.right_stop_btn.setObjectName("stopVideoBtn")
        self.right_stop_btn.clicked.connect(self._stop_video)
        self.right_stop_btn.hide()
        cam_actions.addWidget(self.right_stop_btn)
        self.right_cam_actions = cam_actions
        mode_layout.addLayout(cam_actions)
        # 扫描摄像头填充下拉
        self._scan_right_cameras()
        layout.addWidget(mode_frame)

        # ── 2. 统计卡片（2×2 网格） ──
        stats_frame = QFrame()
        stats_frame.setObjectName("card")
        stats_layout = QGridLayout(stats_frame)
        stats_layout.setContentsMargins(8, 8, 8, 8)
        stats_layout.setSpacing(4)

        self._pass_count_label = QLabel("0")
        pass_card = self._create_mini_card("● 合格", self._pass_count_label, COLOR_OK)
        stats_layout.addWidget(pass_card, 0, 0)

        self._fail_count_label = QLabel("0")
        fail_card = self._create_mini_card("● 不合格", self._fail_count_label, COLOR_ERR)
        stats_layout.addWidget(fail_card, 0, 1)

        self._total_count_label = QLabel("0")
        total_card = self._create_mini_card("● 总数", self._total_count_label, COLOR_ACCENT)
        stats_layout.addWidget(total_card, 1, 0)

        self._rate_label = QLabel("0%")
        rate_card = self._create_mini_card("● 合格率", self._rate_label, COLOR_WARN)
        stats_layout.addWidget(rate_card, 1, 1)

        layout.addWidget(stats_frame)

        # ── 3. 检测结果表格 ──
        table_frame = QFrame()
        table_frame.setObjectName("card")
        table_layout = QVBoxLayout(table_frame)
        table_layout.setContentsMargins(6, 6, 6, 6)
        table_layout.setSpacing(2)
        table_title = QLabel("检测结果")
        table_title.setStyleSheet(f"color: {COLOR_TEXT_DIM}; font-size: 11px; font-weight: 600;")
        table_layout.addWidget(table_title)

        self.result_table = QTableWidget()
        self.result_table.setColumnCount(5)
        self.result_table.setHorizontalHeaderLabels(["#", "类别", "置信度", "严重程度", "状态"])
        self.result_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.result_table.verticalHeader().setVisible(False)
        self.result_table.setAlternatingRowColors(False)
        self.result_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.result_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.result_table.setMinimumHeight(120)
        table_layout.addWidget(self.result_table)
        layout.addWidget(table_frame)

        # ── 4. 缺陷分布 ──
        defect_frame = QFrame()
        defect_frame.setObjectName("card")
        defect_layout = QVBoxLayout(defect_frame)
        defect_layout.setContentsMargins(6, 6, 6, 6)
        defect_layout.setSpacing(2)
        defect_title = QLabel("缺陷分布")
        defect_title.setStyleSheet(f"color: {COLOR_TEXT_DIM}; font-size: 11px; font-weight: 600;")
        defect_layout.addWidget(defect_title)

        self.defect_bars = {}
        for class_name, class_name_cn in zip(CLASS_NAMES, CLASS_NAMES_CN):
            bar_widget = self._create_defect_bar(class_name, class_name_cn)
            defect_layout.addWidget(bar_widget)
            self.defect_bars[class_name] = bar_widget
        layout.addWidget(defect_frame)

        # ── 5. 拍照记录（摄像头模式可见） ──
        self.snapshot_frame = QFrame()
        self.snapshot_frame.setObjectName("card")
        snap_layout = QVBoxLayout(self.snapshot_frame)
        snap_layout.setContentsMargins(6, 6, 6, 6)
        snap_layout.setSpacing(2)
        snap_title = QLabel("拍照记录")
        snap_title.setStyleSheet(f"color: {COLOR_TEXT_DIM}; font-size: 11px; font-weight: 600;")
        snap_layout.addWidget(snap_title)

        self.snapshot_table = QTableWidget()
        self.snapshot_table.setColumnCount(3)
        self.snapshot_table.setHorizontalHeaderLabels(["时间", "判定", "缺陷"])
        self.snapshot_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.snapshot_table.verticalHeader().setVisible(False)
        self.snapshot_table.setAlternatingRowColors(False)
        self.snapshot_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.snapshot_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.snapshot_table.setMinimumHeight(80)
        snap_layout.addWidget(self.snapshot_table)
        layout.addWidget(self.snapshot_frame)

        # 拍照按钮（摄像头模式下显示在面板底部）
        self.right_snapshot_btn = QPushButton("  拍照统计")
        self.right_snapshot_btn.setObjectName("snapshotBtn")
        self.right_snapshot_btn.clicked.connect(self._take_snapshot)
        layout.addWidget(self.right_snapshot_btn)

        # ── 6. 导出 / 设置 / 监控 ──
        actions_row = QHBoxLayout()
        actions_row.setSpacing(4)
        export_btn2 = QPushButton("导出报告")
        export_btn2.clicked.connect(self._export_report)
        actions_row.addWidget(export_btn2)
        import_model_btn2 = QPushButton("导入模型")
        import_model_btn2.clicked.connect(self._import_model)
        actions_row.addWidget(import_model_btn2)
        layout.addLayout(actions_row)

        web_row = QHBoxLayout()
        web_row.setSpacing(4)
        web_btn2 = QPushButton("Web 监控")
        web_btn2.clicked.connect(self._start_web_monitor)
        web_row.addWidget(web_btn2)
        history_btn2 = QPushButton("缺陷记录")
        history_btn2.clicked.connect(self._open_defect_folder)
        web_row.addWidget(history_btn2)
        export_annotated_btn = QPushButton("导出标注图")
        export_annotated_btn.clicked.connect(self._batch_export_annotated)
        web_row.addWidget(export_annotated_btn)
        layout.addLayout(web_row)

        # 导出 6MP 按钮
        export_6mp_btn = QPushButton("  导出 6MP 图片")
        export_6mp_btn.setStyleSheet(f"""
            QPushButton {{
                background: {COLOR_ACCENT};
                color: white;
                border: none;
                padding: 8px;
                font-weight: 600;
            }}
            QPushButton:hover {{
                background: #5aaeff;
            }}
        """)
        export_6mp_btn.clicked.connect(self._export_as_6mp)
        layout.addWidget(export_6mp_btn)

        # 检测数量标签
        self.detection_count_badge = QLabel("0 项检测")
        self.detection_count_badge.setAlignment(Qt.AlignCenter)
        self.detection_count_badge.setStyleSheet(f"""
            background-color: {COLOR_ACCENT}; color: white;
            padding: 4px; border-radius: 8px;
            font-size: 11px; font-weight: 600;
        """)
        layout.addWidget(self.detection_count_badge)

        layout.addStretch()

        scroll.setWidget(panel)
        outer_layout.addWidget(scroll)

        self.right_image_widgets = [open_btn2, open_dir_btn2, self.start_btn2, self.clear_btn2]
        self.right_cam_widgets = [self.right_camera_combo, self.right_camera_btn, self.right_phone_btn, self.right_video_btn, self.right_stop_btn]
        self.right_panel_widgets = {
            "snapshot_frame": self.snapshot_frame,
            "snapshot_btn": self.right_snapshot_btn,
        }

        # 初始状态：图片模式
        self._sync_right_mode_ui(0)

        return outer

    def _create_mini_card(self, title, value_label, color):
        """创建迷你统计卡片"""
        card = QFrame()
        card.setStyleSheet(f"""
            QFrame {{
                background: {COLOR_BG_ALT};
                border: 1px solid {COLOR_BORDER};
                border-left: 2px solid {color};
                border-radius: 4px;
            }}
        """)
        layout = QHBoxLayout(card)
        layout.setContentsMargins(8, 6, 8, 6)
        layout.setSpacing(6)

        title_lbl = QLabel(title)
        title_lbl.setStyleSheet(f"color: {COLOR_TEXT_DIM}; font-size: 11px;")
        layout.addWidget(title_lbl)
        layout.addStretch()
        value_label.setStyleSheet(f"color: {color}; font-size: 16px; font-weight: 700;")
        layout.addWidget(value_label)

        return card

    def _create_stat_inline(self, title, value_label, color):
        """创建行内统计（紧凑横向）"""
        widget = QWidget()
        layout = QHBoxLayout(widget)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(4)

        # 彩色圆点
        dot = QLabel("●")
        dot.setStyleSheet(f"color: {color}; font-size: 8px;")
        layout.addWidget(dot)

        # 标题
        title_label = QLabel(title)
        title_label.setStyleSheet(f"color: {COLOR_TEXT_DIM}; font-size: 11px;")
        layout.addWidget(title_label)

        # 数值
        value_label.setStyleSheet(f"""
            color: {color};
            font-size: 16px;
            font-weight: 700;
        """)
        layout.addWidget(value_label)

        return widget

    def _create_stat_card(self, title, value_label, color, card_type):
        """创建统计卡片（紧凑版）"""
        card = QFrame()
        card.setObjectName("statCard")
        card.setProperty("cardType", card_type)
        card.setStyleSheet(f"""
            QFrame#statCard {{
                background-color: {COLOR_BG_ALT};
                border: 1px solid {COLOR_BORDER};
                border-radius: 8px;
                padding: 8px;
                border-left: 3px solid {color};
            }}
        """)

        layout = QVBoxLayout(card)
        layout.setSpacing(4)
        layout.setContentsMargins(6, 4, 6, 4)

        # 图标 + 标题
        header = QHBoxLayout()
        icon_label = QLabel(title.split()[0])
        icon_label.setStyleSheet(f"font-size: 13px;")
        header.addWidget(icon_label)

        title_label = QLabel(title.split()[-1] if len(title.split()) > 1 else title)
        title_label.setStyleSheet(f"""
            color: {COLOR_TEXT_DIM};
            font-size: 11px;
            font-weight: 500;
        """)
        header.addWidget(title_label)
        header.addStretch()
        layout.addLayout(header)

        # 数值
        value_label.setStyleSheet(f"""
            font-size: 22px;
            font-weight: 700;
            color: {COLOR_TEXT};
        """)
        layout.addWidget(value_label)

        return card

    def _create_defect_bar(self, class_name, class_name_cn):
        """创建缺陷分布条（紧凑版）"""
        # 获取类别对应的颜色
        from constants import CLASS_COLORS_RGB
        color_rgb = CLASS_COLORS_RGB.get(class_name, (74, 158, 255))
        color_hex = f"#{color_rgb[0]:02x}{color_rgb[1]:02x}{color_rgb[2]:02x}"
        
        widget = QWidget()
        layout = QHBoxLayout(widget)
        layout.setContentsMargins(0, 2, 0, 2)
        layout.setSpacing(4)

        # 类别名
        name_label = QLabel(class_name_cn)
        name_label.setFixedWidth(40)
        name_label.setStyleSheet(f"""
            color: {COLOR_TEXT};
            font-size: 11px;
        """)
        layout.addWidget(name_label)

        # 进度条
        progress = QProgressBar()
        progress.setFixedHeight(6)
        progress.setTextVisible(False)
        progress.setStyleSheet(f"""
            QProgressBar {{
                background-color: {COLOR_SURFACE};
                border-radius: 3px;
            }}
            QProgressBar::chunk {{
                background: {color_hex};
                border-radius: 3px;
            }}
        """)
        layout.addWidget(progress)

        # 数量
        count_label = QLabel("0")
        count_label.setFixedWidth(24)
        count_label.setStyleSheet(f"""
            color: {color_hex};
            font-weight: 600;
            font-size: 11px;
        """)
        count_label.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
        layout.addWidget(count_label)

        widget.progress = progress
        widget.count_label = count_label

        return widget

    def _init_menu(self):
        """初始化菜单栏"""
        menubar = self.menuBar()

        # 文件菜单
        file_menu = menubar.addMenu("文件(&F)")

        open_action = QAction("打开图片(&O)", self)
        open_action.setShortcut("Ctrl+O")
        open_action.triggered.connect(self._open_image)
        file_menu.addAction(open_action)

        open_dir_action = QAction("打开文件夹(&D)", self)
        open_dir_action.setShortcut("Ctrl+Shift+O")
        open_dir_action.triggered.connect(self._open_directory)
        file_menu.addAction(open_dir_action)

        file_menu.addSeparator()

        export_action = QAction("导出报告(&E)", self)
        export_action.setShortcut("Ctrl+E")
        export_action.triggered.connect(self._export_report)
        file_menu.addAction(export_action)

        export_6mp_action = QAction("导出 6MP 图片(&6)", self)
        export_6mp_action.setShortcut("Ctrl+6")
        export_6mp_action.triggered.connect(self._export_as_6mp)
        file_menu.addAction(export_6mp_action)

        file_menu.addSeparator()

        # 最近文件子菜单
        self.recent_menu = file_menu.addMenu("最近打开(&R)")
        self._update_recent_files_menu()

        file_menu.addSeparator()

        exit_action = QAction("退出(&X)", self)
        exit_action.setShortcut("Ctrl+Q")
        exit_action.triggered.connect(self.close)
        file_menu.addAction(exit_action)

        # 视频菜单
        video_menu = menubar.addMenu("视频(&V)")

        self.camera_menu_action = QAction("打开摄像头(&C)", self)
        self.camera_menu_action.setShortcut("Ctrl+Shift+C")
        self.camera_menu_action.triggered.connect(self._start_camera)
        video_menu.addAction(self.camera_menu_action)

        self.phone_camera_menu_action = QAction("手机摄像头(&P)", self)
        self.phone_camera_menu_action.setShortcut("Ctrl+Shift+P")
        self.phone_camera_menu_action.triggered.connect(self._start_phone_camera)
        video_menu.addAction(self.phone_camera_menu_action)

        self.video_file_menu_action = QAction("打开视频文件(&F)", self)
        self.video_file_menu_action.setShortcut("Ctrl+Shift+V")
        self.video_file_menu_action.triggered.connect(self._open_video_file)
        video_menu.addAction(self.video_file_menu_action)

        video_menu.addSeparator()

        self.stop_video_menu_action = QAction("停止视频(&S)", self)
        self.stop_video_menu_action.setShortcut("Escape")
        self.stop_video_menu_action.triggered.connect(self._stop_video)
        self.stop_video_menu_action.setEnabled(False)
        video_menu.addAction(self.stop_video_menu_action)

        self.snapshot_menu_action = QAction("拍照(&P)", self)
        self.snapshot_menu_action.setShortcut("Space")
        self.snapshot_menu_action.triggered.connect(self._take_snapshot)
        self.snapshot_menu_action.setEnabled(False)
        video_menu.addAction(self.snapshot_menu_action)

        # 设置菜单
        settings_menu = menubar.addMenu("设置(&S)")

        settings_action = QAction("检测参数(&P)", self)
        settings_action.triggered.connect(self._open_settings)
        settings_menu.addAction(settings_action)

        settings_menu.addSeparator()

        auto_save_action = QAction("自动保存缺陷图", self)
        auto_save_action.setCheckable(True)
        auto_save_action.setChecked(self.settings["auto_save_defects"])
        auto_save_action.triggered.connect(self._toggle_auto_save)
        settings_menu.addAction(auto_save_action)

        # 监控菜单
        monitor_menu = menubar.addMenu("监控(&M)")

        web_action = QAction("启动 Web 监控(&W)", self)
        web_action.setShortcut("Ctrl+W")
        web_action.triggered.connect(self._start_web_monitor)
        monitor_menu.addAction(web_action)

        open_browser_action = QAction("打开浏览器(&B)", self)
        open_browser_action.triggered.connect(self._open_web_browser)
        monitor_menu.addAction(open_browser_action)

        # 帮助菜单
        help_menu = menubar.addMenu("帮助(&H)")

        about_action = QAction("关于(&A)", self)
        about_action.triggered.connect(self._show_about)
        help_menu.addAction(about_action)

    def _init_toolbar(self):
        """初始化工具栏"""
        toolbar = QToolBar()
        toolbar.setMovable(False)
        self.addToolBar(toolbar)

        # ── 检测模式选择 ──
        mode_label = QLabel("  检测模式 ")
        mode_label.setStyleSheet(f"color: {COLOR_TEXT_DIM}; font-size: 12px;")
        toolbar.addWidget(mode_label)

        self.mode_combo = QComboBox()
        self.mode_combo.addItems(["图片检测", "摄像头检测", "视频检测"])
        self.mode_combo.setFixedWidth(120)
        self.mode_combo.currentIndexChanged.connect(self._on_mode_changed)
        toolbar.addWidget(self.mode_combo)

        toolbar.addSeparator()

        # ── 图片模式控件 ──
        self.image_widgets = []
        open_btn = QPushButton("  打开图片")
        open_btn.clicked.connect(self._open_image)
        toolbar.addWidget(open_btn)
        self.image_widgets.append(open_btn)

        open_dir_btn = QPushButton("  打开文件夹")
        open_dir_btn.clicked.connect(self._open_directory)
        toolbar.addWidget(open_dir_btn)
        self.image_widgets.append(open_dir_btn)

        self.start_btn = QPushButton("  开始检测")
        self.start_btn.setObjectName("startBtn")
        self.start_btn.clicked.connect(self._start_detection)
        self.start_btn.setEnabled(False)
        self.start_btn2.setEnabled(False)
        toolbar.addWidget(self.start_btn)
        self.image_widgets.append(self.start_btn)

        self.clear_btn = QPushButton("  清除")
        self.clear_btn.setObjectName("clearBtn")
        self.clear_btn.clicked.connect(self._clear_results)
        toolbar.addWidget(self.clear_btn)
        self.image_widgets.append(self.clear_btn)

        # ── 摄像头/视频模式控件 ──
        self.video_widgets = []

        # 摄像头设备选择
        cam_label = QLabel("  摄像头 ")
        cam_label.setStyleSheet(f"color: {COLOR_TEXT_DIM}; font-size: 12px;")
        toolbar.addWidget(cam_label)
        self.video_widgets.append(cam_label)

        self.camera_combo = QComboBox()
        self.camera_combo.setFixedWidth(180)
        self._scan_cameras()
        self.camera_combo.currentIndexChanged.connect(self._on_camera_changed)
        toolbar.addWidget(self.camera_combo)
        self.video_widgets.append(self.camera_combo)

        # 启动摄像头按钮
        self.camera_btn = QPushButton("  启动摄像头")
        self.camera_btn.setObjectName("cameraBtn")
        self.camera_btn.clicked.connect(self._start_camera)
        toolbar.addWidget(self.camera_btn)
        self.video_widgets.append(self.camera_btn)

        # 手机摄像头按钮
        self.phone_camera_btn = QPushButton("  📱 手机摄像头")
        self.phone_camera_btn.setObjectName("cameraBtn")
        self.phone_camera_btn.setStyleSheet(f"""
            QPushButton {{
                background: {COLOR_ACCENT};
                color: white;
                border: none;
                padding: 6px 12px;
                font-weight: 600;
            }}
            QPushButton:hover {{
                background: #5aaeff;
            }}
        """)
        self.phone_camera_btn.clicked.connect(self._start_phone_camera)
        toolbar.addWidget(self.phone_camera_btn)
        self.video_widgets.append(self.phone_camera_btn)

        # 视频文件按钮
        self.video_file_btn = QPushButton("  视频文件")
        self.video_file_btn.setObjectName("videoBtn")
        self.video_file_btn.clicked.connect(self._open_video_file)
        toolbar.addWidget(self.video_file_btn)
        self.video_widgets.append(self.video_file_btn)

        # 停止视频
        self.stop_video_btn = QPushButton("  停止")
        self.stop_video_btn.setObjectName("stopVideoBtn")
        self.stop_video_btn.clicked.connect(self._stop_video)
        self.stop_video_btn.hide()
        toolbar.addWidget(self.stop_video_btn)
        self.video_widgets.append(self.stop_video_btn)

        # 拍照
        self.snapshot_btn = QPushButton("  拍照")
        self.snapshot_btn.setObjectName("snapshotBtn")
        self.snapshot_btn.clicked.connect(self._take_snapshot)
        self.snapshot_btn.hide()
        toolbar.addWidget(self.snapshot_btn)
        self.video_widgets.append(self.snapshot_btn)

        # FPS
        self.fps_label = QLabel("FPS: --")
        self.fps_label.setObjectName("fpsLabel")
        self.fps_label.hide()
        toolbar.addWidget(self.fps_label)
        self.video_widgets.append(self.fps_label)

        toolbar.addSeparator()

        # ── 通用控件 ──
        # 导出报告
        export_btn = QPushButton("  导出报告")
        export_btn.clicked.connect(self._export_report)
        toolbar.addWidget(export_btn)

        # 导出 6MP
        export_6mp_btn = QPushButton("  导出 6MP")
        export_6mp_btn.setStyleSheet(f"""
            QPushButton {{
                background: {COLOR_ACCENT};
                color: white;
                border: none;
                padding: 6px 12px;
                font-weight: 600;
            }}
            QPushButton:hover {{
                background: #5aaeff;
            }}
        """)
        export_6mp_btn.clicked.connect(self._export_as_6mp)
        toolbar.addWidget(export_6mp_btn)

        # 导入模型
        import_model_btn = QPushButton("  导入模型")
        import_model_btn.clicked.connect(self._import_model)
        toolbar.addWidget(import_model_btn)

        # 置信度阈值
        conf_label = QLabel("  置信度 ")
        conf_label.setStyleSheet(f"color: {COLOR_TEXT_DIM}; font-size: 12px;")
        toolbar.addWidget(conf_label)

        self.conf_spin = QDoubleSpinBox()
        self.conf_spin.setRange(0.05, 0.95)
        self.conf_spin.setSingleStep(0.05)
        self.conf_spin.setValue(self.settings["conf_threshold"])
        self.conf_spin.setFixedWidth(70)
        self.conf_spin.valueChanged.connect(self._on_conf_changed)
        toolbar.addWidget(self.conf_spin)

        # 灰度模式
        self.grayscale_cb = QCheckBox("灰度模式")
        self.grayscale_cb.setStyleSheet(f"color: {COLOR_TEXT_DIM}; font-size: 12px;")
        self.grayscale_cb.toggled.connect(self._on_grayscale_toggled)
        toolbar.addWidget(self.grayscale_cb)

        toolbar.addSeparator()

        # 180秒计时
        self.timer_btn = QPushButton("  180秒计时")
        self.timer_btn.setObjectName("timerBtn")
        self.timer_btn.setFixedSize(100, 32)
        self.timer_btn.clicked.connect(self._toggle_timer_mode)
        toolbar.addWidget(self.timer_btn)

        toolbar.addSeparator()

        # 进度条
        self.progress_bar = QProgressBar()
        self.progress_bar.setMaximumWidth(200)
        self.progress_bar.hide()
        toolbar.addWidget(self.progress_bar)

        # 弹簧
        spacer = QWidget()
        spacer.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        toolbar.addWidget(spacer)

        # 初始状态：图片模式
        self._on_mode_changed(0)

    def _init_statusbar(self):
        """初始化状态栏"""
        status_bar = self.statusBar()

        # 左侧状态
        self.status_label = QLabel("就绪")
        status_bar.addWidget(self.status_label)

        # 右侧永久信息
        self.model_info_label = QLabel("模型: 加载中...")
        status_bar.addPermanentWidget(self.model_info_label)

        self.device_info_label = QLabel("设备: --")
        status_bar.addPermanentWidget(self.device_info_label)

        version_info = QLabel(f"v{APP_VERSION}")
        status_bar.addPermanentWidget(version_info)

    def _connect_signals(self):
        """连接信号"""
        pass

    def _scan_cameras(self):
        """扫描可用摄像头设备（静默模式，抑制 OpenCV 错误输出）"""
        self.camera_combo.clear()
        available = []
        # 临时重定向 stderr 抑制 OpenCV 错误日志
        import ctypes
        libc = ctypes.cdll.msvcrt if os.name == 'nt' else ctypes.CDLL(None)
        devnull = os.open(os.devnull, os.O_WRONLY)
        old_stderr = os.dup(2)
        os.dup2(devnull, 2)
        os.close(devnull)
        try:
            for i in range(5):
                for backend in (cv2.CAP_DSHOW, cv2.CAP_ANY):
                    try:
                        cap = cv2.VideoCapture(i, backend)
                    except Exception:
                        continue
                    if cap.isOpened():
                        name = f"摄像头 {i}"
                        try:
                            w = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
                            h = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
                            if w > 0 and h > 0:
                                name += f" ({w}x{h})"
                        except Exception:
                            pass
                        available.append((i, name))
                        cap.release()
                        break
                    cap.release()
        finally:
            os.dup2(old_stderr, 2)
            os.close(old_stderr)
        if available:
            for idx, name in available:
                self.camera_combo.addItem(name, idx)
        else:
            self.camera_combo.addItem("未检测到摄像头", -1)

    def _on_camera_changed(self, index):
        """摄像头设备切换（工具栏 + 右侧面板同步）"""
        if index >= 0:
            self.right_camera_combo.setCurrentIndex(index)
        if self.is_video_streaming and self.mode_combo.currentIndex() == 1:
            self._stop_video()

    def _on_mode_changed(self, index):
        """检测模式切换：0=图片, 1=摄像头, 2=视频"""
        if self.is_video_streaming:
            self._stop_video()

        is_image = (index == 0)
        is_camera = (index == 1)

        # 工具栏控件
        for w in self.image_widgets:
            w.setVisible(is_image)
        for w in self.video_widgets:
            w.setVisible(not is_image)
        self.camera_combo.setVisible(is_camera)
        self.camera_btn.setVisible(is_camera)
        self.phone_camera_btn.setVisible(is_camera)  # 手机按钮只在摄像头模式显示
        self.video_file_btn.setVisible(index == 2)
        self.stop_video_btn.setVisible(False)
        self.snapshot_btn.setVisible(False)
        self.fps_label.setVisible(False)
        self.start_btn.setEnabled(is_image and bool(self.current_image_path))
        self.start_btn2.setEnabled(is_image and bool(self.current_image_path))
        self.camera_menu_action.setEnabled(not is_image)
        self.phone_camera_menu_action.setEnabled(not is_image)
        self.video_file_menu_action.setEnabled(not is_image)

        # 同步右侧面板
        self.right_mode_combo.blockSignals(True)
        self.right_mode_combo.setCurrentIndex(index)
        self.right_mode_combo.blockSignals(False)
        self._sync_right_mode_ui(index)

    def _scan_right_cameras(self):
        """扫描摄像头填充右侧面板下拉框（静默模式）"""
        self.right_camera_combo.clear()
        available = []
        # 临时重定向 stderr 抑制 OpenCV 错误日志
        devnull = os.open(os.devnull, os.O_WRONLY)
        old_stderr = os.dup(2)
        os.dup2(devnull, 2)
        os.close(devnull)
        try:
            for i in range(5):
                for backend in (cv2.CAP_DSHOW, cv2.CAP_ANY):
                    try:
                        cap = cv2.VideoCapture(i, backend)
                    except Exception:
                        continue
                    if cap.isOpened():
                        name = f"摄像头 {i}"
                        try:
                            w = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
                            h = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
                            if w > 0 and h > 0:
                                name += f" ({w}x{h})"
                        except Exception:
                            pass
                        available.append((i, name))
                        cap.release()
                        break
                    cap.release()
        finally:
            os.dup2(old_stderr, 2)
            os.close(old_stderr)
        if available:
            for idx, name in available:
                self.right_camera_combo.addItem(name, idx)
        else:
            self.right_camera_combo.addItem("未检测到摄像头", -1)

    def _on_right_mode_changed(self, index):
        """右侧面板模式切换 → 同步工具栏"""
        self.mode_combo.blockSignals(True)
        self.mode_combo.setCurrentIndex(index)
        self.mode_combo.blockSignals(False)
        self._on_mode_changed(index)

    def _sync_right_mode_ui(self, index):
        """同步右侧面板 UI 到指定模式"""
        is_image = (index == 0)
        is_camera = (index == 1)
        is_video = (index == 2)

        for w in self.right_image_widgets:
            w.setVisible(is_image)
        for w in self.right_cam_widgets:
            w.setVisible(not is_image)
        self.right_camera_combo.setVisible(is_camera)
        self.right_camera_btn.setVisible(is_camera)
        self.right_phone_btn.setVisible(is_camera)  # 手机按钮只在摄像头模式显示
        self.right_video_btn.setVisible(is_video)
        self.right_stop_btn.setVisible(False)
        self.right_panel_widgets["snapshot_frame"].setVisible(is_camera)
        self.right_panel_widgets["snapshot_btn"].setVisible(is_camera)

    def _init_detector(self):
        """初始化检测器"""
        try:
            from inference import ScrewDefectDetector, find_best_model

            # 查找模型
            model_path = self.settings["model_path"]
            if not model_path or not Path(model_path).exists():
                # 打包后的路径
                if getattr(sys, 'frozen', False):
                    base_path = Path(sys._MEIPASS)
                    packaged_model = base_path / "best.pt"
                    if packaged_model.exists():
                        model_path = str(packaged_model)
                    else:
                        model_path = "yolov8n.pt"
                else:
                    best = find_best_model()
                    if best:
                        model_path = str(best)
                    else:
                        model_path = "yolov8n.pt"

            self.status_label.setText(f"正在加载模型: {Path(model_path).name}...")

            # 创建检测器
            self.detector = ScrewDefectDetector(
                model_path=model_path,
                conf_threshold=self.settings["conf_threshold"],
                iou_threshold=self.settings["iou_threshold"],
                img_size=self.settings["img_size"],
            )

            # 更新状态栏
            device_info = self.detector.get_device_info()
            self.model_info_label.setText(f"模型: {Path(model_path).name}")
            self.device_info_label.setText(f"设备: {device_info['device'].upper()}")

            self.status_label.setText("模型加载完成")

        except Exception as e:
            self.status_label.setText(f"模型加载失败: {str(e)}")
            QMessageBox.warning(
                self, "警告",
                f"模型加载失败:\n{str(e)}\n\n请检查模型文件是否存在。"
            )

    def _on_conf_changed(self, value):
        """置信度阈值变更"""
        self.settings["conf_threshold"] = value
        if self.detector:
            self.detector.conf_threshold = value

    def _on_grayscale_toggled(self, checked):
        """灰度模式切换 — 同步工具栏和右侧面板"""
        self.grayscale_enabled = checked
        # 同步两个复选框（避免信号循环）
        self.grayscale_cb.blockSignals(True)
        self.right_grayscale_cb.blockSignals(True)
        self.grayscale_cb.setChecked(checked)
        self.right_grayscale_cb.setChecked(checked)
        self.grayscale_cb.blockSignals(False)
        self.right_grayscale_cb.blockSignals(False)
        # 如果当前有图片，刷新显示
        if self.current_image_path and Path(self.current_image_path).is_file():
            self._display_image(self.current_image_path)
        status = "已开启" if checked else "已关闭"
        self.status_label.setText(f"灰度模式: {status}")

    def _apply_grayscale(self, frame):
        """将 BGR 帧转换为灰度（3通道），用于显示和检测"""
        if not self.grayscale_enabled:
            return frame
        gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        return cv2.cvtColor(gray, cv2.COLOR_GRAY2BGR)

    def _import_model(self):
        """导入自定义模型文件"""
        file_path, _ = QFileDialog.getOpenFileName(
            self, "选择模型文件", "",
            "模型文件 (*.pt *.onnx);;所有文件 (*)"
        )
        if not file_path:
            return

        if not Path(file_path).exists():
            QMessageBox.warning(self, "错误", "模型文件不存在。")
            return

        try:
            from inference import ScrewDefectDetector

            self.status_label.setText(f"正在加载模型: {Path(file_path).name}...")

            # 创建新的检测器
            new_detector = ScrewDefectDetector(
                model_path=file_path,
                conf_threshold=self.settings["conf_threshold"],
                iou_threshold=self.settings["iou_threshold"],
                img_size=self.settings["img_size"],
            )

            # 替换旧检测器
            self.detector = new_detector
            self.settings["model_path"] = file_path

            # 更新界面信息
            device_info = self.detector.get_device_info()
            self.model_info_label.setText(f"模型: {Path(file_path).name}")
            self.device_info_label.setText(f"设备: {device_info['device'].upper()}")
            self.status_label.setText(f"模型加载成功: {Path(file_path).name}")

        except Exception as e:
            QMessageBox.warning(
                self, "加载失败",
                f"模型加载失败:\n{str(e)}"
            )
            self.status_label.setText("模型加载失败")

    # ── 以下为功能方法 ──

    def _update_recent_files_menu(self):
        """更新最近文件菜单"""
        self.recent_menu.clear()
        try:
            recent = db.get_recent_files(10)
            if not recent:
                action = QAction("无记录", self)
                action.setEnabled(False)
                self.recent_menu.addAction(action)
            else:
                for item in recent:
                    fp = item['file_path']
                    name = Path(fp).name
                    action = QAction(f"{name}", self)
                    action.setToolTip(fp)
                    action.triggered.connect(lambda checked, path=fp: self._open_recent_file(path))
                    self.recent_menu.addAction(action)

                self.recent_menu.addSeparator()
                clear_action = QAction("清空记录", self)
                clear_action.triggered.connect(self._clear_recent_files)
                self.recent_menu.addAction(clear_action)
        except Exception:
            pass

    def _open_recent_file(self, file_path):
        """打开最近文件"""
        if Path(file_path).exists():
            self.image_list = []
            self.current_image_index = -1
            self.current_image_path = file_path
            self._display_image(file_path)
            self.start_btn.setEnabled(True)
            self.start_btn2.setEnabled(True)
            self.prev_btn.setEnabled(False)
            self.next_btn.setEnabled(False)
            self.image_info_label.setText(f"文件: {Path(file_path).name}")
            self.status_label.setText(f"已加载: {Path(file_path).name}")
        else:
            QMessageBox.warning(self, "警告", f"文件不存在:\n{file_path}")

    def _clear_recent_files(self):
        """清空最近文件记录"""
        db.clear_recent_files()
        self._update_recent_files_menu()

    def _open_image(self):
        """打开单张图片"""
        if self.is_video_streaming or self.is_detecting:
            return

        file_path, _ = QFileDialog.getOpenFileName(
            self, "选择图片", "",
            "图片文件 (*.jpg *.jpeg *.png *.bmp *.tiff);;所有文件 (*)"
        )

        if file_path:
            self.image_list = []
            self.current_image_index = -1
            self.current_image_path = file_path
            self._display_image(file_path)
            self.start_btn.setEnabled(True)
            self.start_btn2.setEnabled(True)
            self.prev_btn.setEnabled(False)
            self.next_btn.setEnabled(False)
            self.image_info_label.setText(f"文件: {Path(file_path).name}")
            self.status_label.setText(f"已加载: {Path(file_path).name}")
            # 记录到最近文件
            db.add_recent_file(file_path, "image")
            self._update_recent_files_menu()

    def _open_directory(self):
        """打开文件夹"""
        if self.is_video_streaming or self.is_detecting:
            return

        dir_path = QFileDialog.getExistingDirectory(self, "选择图片文件夹")
        if dir_path:
            # 记录到最近文件
            db.add_recent_file(dir_path, "directory")
            self.current_image_path = dir_path
            extensions = ['*.jpg', '*.jpeg', '*.png', '*.bmp', '*.tiff']
            image_files = set()
            for ext in extensions:
                image_files.update(Path(dir_path).glob(ext))
                image_files.update(Path(dir_path).glob(ext.upper()))

            if image_files:
                sorted_files = sorted(image_files)
                self.image_list = [str(p) for p in sorted_files]
                self.current_image_index = 0

                self.start_btn.setEnabled(True)
                self.start_btn2.setEnabled(True)
                self._display_image(self.image_list[0])
                self._update_nav_buttons()
                self.image_info_label.setText(
                    f"1/{len(self.image_list)}  {Path(self.image_list[0]).name}"
                )
                self.status_label.setText(f"已加载 {len(self.image_list)} 张图片")
            else:
                QMessageBox.information(self, "提示", "该文件夹中没有找到图片文件")

    def _fit_image(self, pixmap, fast=False):
        """将图片缩放适配显示区（完整显示，不裁剪）"""
        if pixmap.isNull():
            return
        # 视频模式使用 FastTransformation 减少卡顿
        transform_mode = Qt.FastTransformation if fast else Qt.SmoothTransformation
        scaled = pixmap.scaled(
            self.image_display.size(),
            Qt.KeepAspectRatio,
            transform_mode
        )
        self.image_display.setPixmap(scaled)

    def _crop_and_resize_to_6mp(self, image_path):
        """自选区缩放到 6MP (3072×2048)"""
        img = cv2.imread(image_path)
        if img is None:
            return None

        h, w = img.shape[:2]
        target_w, target_h = 3072, 2048  # 6MP

        # 如果图片已经小于6MP，直接返回
        if w <= target_w and h <= target_h:
            return img

        # 计算裁剪区域（居中裁剪，保持比例）
        target_ratio = target_w / target_h  # 1.5
        current_ratio = w / h

        if current_ratio > target_ratio:
            # 图片更宽，裁剪宽度
            new_w = int(h * target_ratio)
            x1 = (w - new_w) // 2
            crop = img[:, x1:x1+new_w]
        else:
            # 图片更高，裁剪高度
            new_h = int(w / target_ratio)
            y1 = (h - new_h) // 2
            crop = img[y1:y1+new_h, :]

        # 缩放到目标尺寸
        resized = cv2.resize(crop, (target_w, target_h), interpolation=cv2.INTER_LANCZOS4)
        return resized

    def _export_as_6mp(self):
        """导出当前图片为 6MP 格式"""
        if not self.current_image_path or not Path(self.current_image_path).is_file():
            QMessageBox.information(self, "提示", "请先打开一张图片")
            return

        save_path, _ = QFileDialog.getSaveFileName(
            self, "导出 6MP 图片", "",
            "JPEG 文件 (*.jpg);;PNG 文件 (*.png)"
        )
        if not save_path:
            return

        try:
            result = self._crop_and_resize_to_6mp(self.current_image_path)
            if result is not None:
                cv2.imwrite(save_path, result)
                QMessageBox.information(self, "成功",
                    f"已导出 6MP 图片:\n{save_path}\n\n"
                    f"分辨率: 3072×2048 (6MP)")
            else:
                QMessageBox.warning(self, "错误", "图片处理失败")
        except Exception as e:
            QMessageBox.critical(self, "错误", f"导出失败: {str(e)}")

    def _display_image(self, image_path):
        """显示图片（支持灰度模式）"""
        if self.grayscale_enabled:
            img = cv2.imread(image_path)
            if img is not None:
                img = self._apply_grayscale(img)
                rgb_img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
                h, w, ch = rgb_img.shape
                q_img = QImage(rgb_img.data, w, h, ch * w, QImage.Format_RGB888).copy()
                pixmap = QPixmap.fromImage(q_img)
            else:
                pixmap = QPixmap(image_path)
        else:
            pixmap = QPixmap(image_path)
        self._fit_image(pixmap)

    def _show_image_at_index(self, index):
        """显示指定索引的图片"""
        if not self.image_list or index < 0 or index >= len(self.image_list):
            return
        self.current_image_index = index
        image_path = self.image_list[index]
        self.current_image_path = image_path
        self._display_image(image_path)
        self._update_nav_buttons()
        self.image_info_label.setText(
            f"{index + 1}/{len(self.image_list)}  {Path(image_path).name}"
        )
        self.status_label.setText(f"已加载: {Path(image_path).name}")

    def _prev_image(self):
        """显示上一张图片"""
        if self.current_image_index > 0:
            self._show_image_at_index(self.current_image_index - 1)

    def _next_image(self):
        """显示下一张图片"""
        if self.current_image_index < len(self.image_list) - 1:
            self._show_image_at_index(self.current_image_index + 1)

    def _update_nav_buttons(self):
        """更新导航按钮状态"""
        has_list = len(self.image_list) > 0
        self.prev_btn.setEnabled(has_list and self.current_image_index > 0)
        self.next_btn.setEnabled(
            has_list and self.current_image_index < len(self.image_list) - 1
        )

    def keyPressEvent(self, event):
        """键盘快捷键"""
        if event.key() == Qt.Key_Left:
            self._prev_image()
        elif event.key() == Qt.Key_Right:
            self._next_image()
        else:
            super().keyPressEvent(event)

    def _start_detection(self):
        """开始检测"""
        if self.is_detecting:
            return

        if not self.current_image_path or not self.detector:
            return

        if self.is_video_streaming:
            return

        self.is_detecting = True
        self.start_btn.setEnabled(False)
        self.start_btn2.setEnabled(False)
        self.status_label.setText("正在检测...")

        path = Path(self.current_image_path)
        if path.is_dir():
            # 批量检测
            self._start_batch_detection(str(path))
        else:
            # 单张检测
            self._detect_single_image(str(path))

    def _detect_single_image(self, image_path):
        """检测单张图片"""
        try:
            # 灰度模式：读取图片并转灰度后送入检测器
            if self.grayscale_enabled:
                img = cv2.imread(image_path)
                if img is not None:
                    gray_img = self._apply_grayscale(img)
                    result = self.detector.detect_single(gray_img)
                    annotated = self.detector.visualize_result(gray_img, result)
                else:
                    result = self.detector.detect_single(image_path)
                    annotated = self.detector.visualize_result(image_path, result)
            else:
                result = self.detector.detect_single(image_path)
                annotated = self.detector.visualize_result(image_path, result)

            # 显示结果
            h, w, ch = annotated.shape
            bytes_per_line = ch * w
            rgb_img = cv2.cvtColor(annotated, cv2.COLOR_BGR2RGB)
            q_img = QImage(rgb_img.data, w, h, bytes_per_line, QImage.Format_RGB888).copy()
            pixmap = QPixmap.fromImage(q_img)
            self._fit_image(pixmap)
            self.all_results.append(result)
            self._update_result_table(result)
            self._update_stats()
            self._update_defect_distribution()

            # 自动保存标注图片（所有图片都保存，不仅仅是缺陷）
            annotated_path = self._auto_save_annotated(image_path, annotated, result)

            # 保存到数据库 + 推送 Web（包含标注图路径）
            self._save_and_push(image_path, result, annotated_path)

            # 反馈机制：语音播报 + 日志记录 + Web 推送
            self.feedback.voice.speak_result(result)
            if self.feedback.logger:
                self.feedback.logger.log_result(result)
            self._push_feedback_web(result)

            # 低置信度人工复核
            self._check_low_confidence_review(image_path, result)

            self.image_info_label.setText(
                f"推理时间: {result.inference_time_ms:.1f}ms | {result.overall_verdict}"
            )
            self.status_label.setText(f"检测完成: {result.overall_verdict}")

        except Exception as e:
            self.status_label.setText(f"检测失败: {str(e)}")

        finally:
            self.is_detecting = False
            self.start_btn.setEnabled(True)
            self.start_btn2.setEnabled(True)

    def _check_low_confidence_review(self, image_path, result):
        """低置信度检测结果人工复核"""
        low_conf_threshold = 0.5
        low_conf_dets = [d for d in result.detections if d.confidence < low_conf_threshold]
        if not low_conf_dets:
            return

        dialog = ReviewDialog(image_path, low_conf_dets, self)
        if dialog.exec_() == QDialog.Accepted and dialog.result_data:
            for det, reviewed_en, reviewed_cn, action in dialog.result_data:
                db.save_review(
                    image_path=image_path,
                    original_class=det.class_name,
                    original_class_cn=det.class_name_cn,
                    original_confidence=det.confidence,
                    reviewed_class=reviewed_en,
                    reviewed_class_cn=reviewed_cn,
                    action=action,
                )
            # 语音反馈
            corrected = sum(1 for _, _, _, a in dialog.result_data if a == "修正")
            if corrected > 0:
                self.feedback.voice.speak_async(f"复核完成，修正{corrected}项")
            else:
                self.feedback.voice.speak_async("复核完成，全部确认")
            self.status_label.setText(f"人工复核完成: {len(dialog.result_data)} 项")

    def _save_and_push(self, image_path, result, annotated_path=None):
        """保存检测结果到数据库并推送到 Web 面板"""
        try:
            det_dicts = []
            for det in result.detections:
                det_dicts.append({
                    'class_name': det.class_name,
                    'class_name_cn': det.class_name_cn,
                    'confidence': det.confidence,
                    'bbox': list(det.bbox),
                    'is_defect': det.is_defect,
                })
            db.save_detection(
                image_path=image_path,
                inference_time_ms=result.inference_time_ms,
                source_type='image',
                annotated_path=annotated_path,
                detections=result.detections,
            )
            # 推送到 Web
            self._push_web({
                'image_path': image_path,
                'annotated_path': annotated_path or '',
                'inference_time_ms': result.inference_time_ms,
                'source_type': 'image',
                'detections': det_dicts,
            })
        except Exception as e:
            print(f"[保存失败] {image_path}: {e}")

    def _push_web(self, data):
        """推送检测结果到 Web 监控面板"""
        try:
            import urllib.request
            port = self.settings.get("web_port", 5000)
            req = urllib.request.Request(
                f"http://127.0.0.1:{port}/api/push",
                data=json.dumps(data, ensure_ascii=False).encode('utf-8'),
                headers={'Content-Type': 'application/json'},
                method='POST'
            )
            urllib.request.urlopen(req, timeout=1)
        except Exception:
            pass

    def _push_feedback_web(self, result):
        """推送反馈事件到 Web 监控面板"""
        try:
            import urllib.request
            port = self.settings.get("web_port", 5000)
            # 构造语音播报文本
            voice_text = ""
            if not result.detections:
                voice_text = "未检测到目标"
            elif not result.has_defect:
                voice_text = "检测合格"
            else:
                parts = ["检测不合格"]
                defect_summary = {}
                for det in result.detections:
                    if det.is_defect:
                        defect_summary[det.class_name_cn] = defect_summary.get(det.class_name_cn, 0) + 1
                for name, count in defect_summary.items():
                    parts.append(f"{count}处{name}")
                voice_text = "，".join(parts)

            feedback_data = {
                'type': 'feedback',
                'timestamp': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                'image_path': result.image_path,
                'has_defect': result.has_defect,
                'verdict': result.overall_verdict,
                'voice_text': voice_text,
                'voice_enabled': self.settings.get("enable_voice", False),
                'log_enabled': self.settings.get("enable_log", True),
                'log_path': str(self.feedback.logger.log_path) if self.feedback.logger else '',
                'inference_time_ms': result.inference_time_ms,
                'defect_count': result.defect_count,
                'detections': [
                    {'class_name_cn': d.class_name_cn, 'confidence': d.confidence, 'is_defect': d.is_defect}
                    for d in result.detections
                ],
            }
            req = urllib.request.Request(
                f"http://127.0.0.1:{port}/api/feedback",
                data=json.dumps(feedback_data, ensure_ascii=False).encode('utf-8'),
                headers={'Content-Type': 'application/json'},
                method='POST'
            )
            urllib.request.urlopen(req, timeout=1)
        except Exception:
            pass

    def _auto_save_defect(self, image_path, annotated_img, result):
        """自动保存缺陷图片到指定目录（单张检测用，保存标注图）"""
        try:
            save_dir = Path(self.settings["auto_save_dir"])
            save_dir.mkdir(parents=True, exist_ok=True)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            defect_types = "_".join(set(d.class_name for d in result.detections))
            filename = f"{timestamp}_{defect_types}.jpg"
            save_path = save_dir / filename
            cv2.imwrite(str(save_path), annotated_img)
        except Exception:
            pass

    def _auto_save_annotated(self, image_path, annotated_img, result):
        """自动保存标注图片（所有检测图片都保存），返回保存路径"""
        try:
            # 创建保存目录
            save_dir = Path(self.settings["auto_save_dir"]) / "annotated"
            save_dir.mkdir(parents=True, exist_ok=True)

            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            original_name = Path(image_path).stem

            # 判定结果
            verdict = "合格" if not result.has_defect else "缺陷"

            # 缺陷类型
            if result.detections:
                defect_types = "_".join(set(d.class_name for d in result.detections))
            else:
                defect_types = "pass"

            # 生成文件名
            filename = f"{timestamp}_{verdict}_{defect_types}_{original_name}.jpg"
            save_path = save_dir / filename

            # 保存标注图片
            cv2.imwrite(str(save_path), annotated_img)

            # 更新状态栏
            self.status_label.setText(
                f"检测完成: {verdict} | 已保存: {Path(save_path).name}"
            )

            return str(save_path)

        except Exception as e:
            print(f"[保存标注图失败] {e}")
            return None

    def _auto_save_defect_copy(self, image_path, result):
        """批量检测时保存缺陷图（直接复制原图，避免内存溢出）"""
        try:
            save_dir = Path(self.settings["auto_save_dir"])
            save_dir.mkdir(parents=True, exist_ok=True)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            defect_types = "_".join(set(d.class_name for d in result.detections))
            ext = Path(image_path).suffix or ".jpg"
            filename = f"{timestamp}_{defect_types}{ext}"
            save_path = save_dir / filename
            shutil.copy2(str(image_path), str(save_path))
        except Exception:
            pass

    def _start_batch_detection(self, dir_path):
        """批量检测"""
        extensions = ['*.jpg', '*.jpeg', '*.png', '*.bmp', '*.tiff']
        image_files = set()
        for ext in extensions:
            image_files.update(Path(dir_path).glob(ext))
            image_files.update(Path(dir_path).glob(ext.upper()))

        image_paths = [str(p) for p in image_files]

        if not image_paths:
            self.status_label.setText("没有找到图片文件")
            self.is_detecting = False
            self.start_btn.setEnabled(True)
            self.start_btn2.setEnabled(True)
            return

        # 创建后台线程
        self.detection_thread = DetectionThread(
            self.detector, image_paths, self,
            grayscale_enabled=self.grayscale_enabled,
        )
        self.detection_thread.progress.connect(self._on_batch_progress)
        self.detection_thread.result_ready.connect(self._on_batch_result)
        self.detection_thread.finished_all.connect(self._on_batch_finished)
        self.detection_thread.error_occurred.connect(self._on_batch_error)

        self.progress_bar.setMaximum(len(image_paths))
        self.progress_bar.setValue(0)
        self.progress_bar.show()

        # 启动批量队列定时器（每 200ms 刷新一次）
        self._batch_queue = []
        self._batch_timer.start(200)

        self.detection_thread.start()

    def _on_batch_progress(self, current, total):
        """批量检测进度"""
        self.progress_bar.setValue(current)
        self.status_label.setText(f"正在检测: {current}/{total}")

    def _on_batch_result(self, image_path, result):
        """批量检测单个结果 — 只入队，不阻塞主线程"""
        self._batch_queue.append((image_path, result))

    def _flush_batch_queue(self):
        """定时刷新：批量处理队列中的结果"""
        if not self._batch_queue:
            return

        queue = self._batch_queue
        self._batch_queue = []

        # UI 更新（轻量，主线程做）
        for image_path, result in queue:
            self.all_results.append(result)
            # 日志记录
            if self.feedback.logger:
                self.feedback.logger.log_result(result)

        # 限制 UI 更新频率，避免频繁重绘
        current_time = time.time()
        if not hasattr(self, '_last_ui_update') or current_time - self._last_ui_update > 0.5:
            self._update_stats()
            self._update_defect_distribution()
            self._last_ui_update = current_time

        total = len(self.all_results)
        self.status_label.setText(f"正在检测... 已处理 {total} 张")

        # 保存和复制扔到后台线程（I/O 密集，不阻塞 UI）
        threading.Thread(
            target=self._batch_save_worker,
            args=(queue,),
            daemon=True,
        ).start()

    def _batch_save_worker(self, queue):
        """后台线程：批量保存数据库 + 生成标注图 + 复制缺陷图"""
        for image_path, result in queue:
            try:
                # 生成标注图
                annotated = self.detector.visualize_result(image_path, result)
                annotated_path = self._auto_save_annotated(image_path, annotated, result)
                
                # 保存到数据库并推送 Web
                self._save_and_push(image_path, result, annotated_path)
            except Exception as e:
                print(f"[批量保存失败] {image_path}: {e}")
                # 即使标注图生成失败，也尝试保存原始结果
                try:
                    self._save_and_push(image_path, result, None)
                except Exception:
                    pass
            
            if self.settings["auto_save_defects"] and result.has_defect:
                try:
                    self._auto_save_defect_copy(image_path, result)
                except Exception:
                    pass
            # 添加小延迟，避免 I/O 过载
            time.sleep(0.01)

    def _on_batch_finished(self, last_result):
        """批量检测完成"""
        # 停止定时器，刷完剩余队列
        self._batch_timer.stop()
        self._flush_batch_queue()

        self.progress_bar.hide()
        self.is_detecting = False
        self.start_btn.setEnabled(True)
        self.start_btn2.setEnabled(True)

        # 最终刷新统计
        self._update_stats()
        self._update_defect_distribution()

        # 显示最后一张结果
        if last_result:
            annotated = self.detector.visualize_result(last_result.image_path, last_result)
            h, w, ch = annotated.shape
            bytes_per_line = ch * w
            rgb_img = cv2.cvtColor(annotated, cv2.COLOR_BGR2RGB)
            q_img = QImage(rgb_img.data, w, h, bytes_per_line, QImage.Format_RGB888).copy()
            pixmap = QPixmap.fromImage(q_img)
            self._fit_image(pixmap)

        total = len(self.all_results)
        pass_count = sum(1 for r in self.all_results if not r.has_defect)
        self.status_label.setText(f"批量检测完成: {total} 张, 合格 {pass_count} 张")

        # 语音播报批量汇总
        self.feedback.voice.speak_batch_summary(total, pass_count, total - pass_count)
        # 日志汇总报告
        if self.feedback.logger:
            self.feedback.logger.log_summary(list(self.all_results))

    def _on_batch_error(self, error_msg):
        """批量检测错误"""
        self.status_label.setText(f"检测错误: {error_msg}")

    def _update_result_table(self, result):
        """更新结果表格（优化版：抑制中间重绘）"""
        # 暂停表格重绘，避免逐单元格触发重绘
        self.result_table.setUpdatesEnabled(False)
        try:
            self.result_table.setRowCount(len(result.detections))

            for i, det in enumerate(result.detections):
                severity, sev_score = db.get_severity(det.class_name, det.confidence)
                sev_color = db.severity_color(severity)
                
                # 获取缺陷类型对应的颜色
                from constants import CLASS_COLORS_RGB
                color_rgb = CLASS_COLORS_RGB.get(det.class_name, (74, 158, 255))
                defect_color = f"#{color_rgb[0]:02x}{color_rgb[1]:02x}{color_rgb[2]:02x}"
                
                for col, text in [
                    (0, str(i + 1)),
                    (1, det.class_name_cn),
                    (2, f"{det.confidence:.2%}"),
                    (3, severity),
                    (4, "缺陷" if det.is_defect else "合格"),
                ]:
                    item = QTableWidgetItem(text)
                    item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled)
                    if col == 3:
                        item.setForeground(QColor(sev_color))
                    elif col == 1 and det.is_defect:
                        # 缺陷类型使用对应的颜色
                        item.setForeground(QColor(defect_color))
                    elif col == 4 and det.is_defect:
                        item.setForeground(QColor(COLOR_ERR))
                    elif col in (1, 4):
                        item.setForeground(QColor(COLOR_OK))
                    self.result_table.setItem(i, col, item)

            self.detection_count_badge.setText(f"{len(result.detections)} 项检测")
        finally:
            # 恢复重绘并强制刷新一次
            self.result_table.setUpdatesEnabled(True)

    def _update_stats(self):
        """更新统计信息"""
        if not self.all_results:
            return

        total = len(self.all_results)
        pass_count = sum(1 for r in self.all_results if not r.has_defect)
        fail_count = total - pass_count
        rate = (pass_count / total * 100) if total > 0 else 0

        self._pass_count_label.setText(str(pass_count))
        self._fail_count_label.setText(str(fail_count))
        self._total_count_label.setText(str(total))
        self._rate_label.setText(f"{rate:.1f}%")

    def _update_defect_distribution(self):
        """更新缺陷分布（优化版：使用缓存计数）"""
        # 初始化缓存计数器
        if not hasattr(self, '_defect_counts_cache'):
            self._defect_counts_cache = {name: 0 for name in CLASS_NAMES}
        
        # 只统计新增的结果
        if not hasattr(self, '_last_result_count'):
            self._last_result_count = 0
        
        # 统计新增的检测结果
        for result in self.all_results[self._last_result_count:]:
            for det in result.detections:
                if det.class_name in self._defect_counts_cache:
                    self._defect_counts_cache[det.class_name] += 1
        
        self._last_result_count = len(self.all_results)
        
        # 找到最大值用于归一化
        max_count = max(self._defect_counts_cache.values()) if self._defect_counts_cache.values() else 1
        
        # 更新进度条
        for class_name, count in self._defect_counts_cache.items():
            if class_name in self.defect_bars:
                widget = self.defect_bars[class_name]
                percent = int((count / max_count) * 100) if max_count > 0 else 0
                widget.progress.setValue(percent)
                widget.count_label.setText(str(count))

    def _clear_results(self):
        """清除结果"""
        if self.is_detecting:
            return
        self.all_results = []
        self.current_result = None
        self.current_image_path = ""
        self.image_list = []
        self.current_image_index = -1
        
        # 重置缓存计数器
        if hasattr(self, '_defect_counts_cache'):
            self._defect_counts_cache = {name: 0 for name in CLASS_NAMES}
        if hasattr(self, '_last_result_count'):
            self._last_result_count = 0

        self.result_table.setRowCount(0)
        self.detection_count_badge.setText("0 项检测")

        self._pass_count_label.setText("0")
        self._fail_count_label.setText("0")
        self._total_count_label.setText("0")
        self._rate_label.setText("0%")

        for class_name in self.defect_bars:
            widget = self.defect_bars[class_name]
            widget.progress.setValue(0)
            widget.count_label.setText("0")

        self.snapshot_table.setRowCount(0)

        self.image_display.clear()
        self.image_display.setText("拖拽图片到此处\n或通过工具栏上传")
        self.image_info_label.setText("")

        self.prev_btn.setEnabled(False)
        self.next_btn.setEnabled(False)
        self.start_btn.setEnabled(False)
        self.start_btn2.setEnabled(False)
        self.status_label.setText("已清除")

    def _start_camera(self):
        """启动摄像头"""
        if self.is_video_streaming:
            return
        if not self.detector:
            QMessageBox.warning(self, "警告", "检测器未初始化")
            return
        # 优先右侧面板选择的摄像头，其次工具栏
        cam_idx = self.right_camera_combo.currentData()
        if cam_idx is None or cam_idx < 0:
            cam_idx = self.camera_combo.currentData()
        if cam_idx is None or cam_idx < 0:
            QMessageBox.warning(self, "警告", "未检测到可用摄像头")
            return
        self._start_video_stream(cam_idx)

    def _open_video_file(self):
        """打开视频文件"""
        if self.is_video_streaming:
            return
        if not self.detector:
            QMessageBox.warning(self, "警告", "检测器未初始化")
            return

        file_path, _ = QFileDialog.getOpenFileName(
            self, "选择视频文件", "",
            "视频文件 (*.mp4 *.avi *.mov *.mkv *.wmv *.flv);;所有文件 (*)"
        )
        if file_path:
            # 记录到最近文件
            db.add_recent_file(file_path, "video")
            self._update_recent_files_menu()
            self._start_video_stream(file_path)

    def _start_phone_camera(self):
        """启动手机摄像头视频流"""
        if self.is_video_streaming:
            return
        if not self.detector:
            QMessageBox.warning(self, "警告", "检测器未初始化")
            return

        dialog = PhoneStreamDialog(self)
        if dialog.exec_() != QDialog.Accepted:
            return

        url = dialog.get_url()
        if not url:
            QMessageBox.warning(self, "错误", "请输入视频流地址")
            return

        # 保存分辨率限制设置
        self._phone_resolution_limit = dialog.get_resolution_limit()

        # 启动手机视频流
        self._start_video_stream(url, is_phone=True)

    def _start_video_stream(self, source, is_phone=False):
        """启动视频流"""
        if self.is_video_streaming:
            return

        self.all_results = []
        self._update_stats()
        self.result_table.setRowCount(0)

        self.is_video_streaming = True
        self._set_video_ui_state(True)

        if isinstance(source, int):
            source_name = f"摄像头 (camera {source})"
        elif is_phone:
            source_name = f"手机摄像头: {source}"
        else:
            source_name = Path(source).name

        self.image_info_label.setText(f"视频流: {source_name}")
        self.status_label.setText(f"正在打开视频流: {source_name}...")

        # 获取分辨率限制
        resolution_limit = getattr(self, '_phone_resolution_limit', None) if is_phone else None

        self.video_thread = VideoStreamThread(
            self.detector, source, self,
            resolution_limit=resolution_limit,
            grayscale_enabled=self.grayscale_enabled,
        )
        self.video_thread.frame_ready.connect(self._on_video_frame)
        self.video_thread.fps_update.connect(self._on_fps_update)
        self.video_thread.error_occurred.connect(self._on_video_error)
        self.video_thread.stream_finished.connect(self._on_video_stream_finished)
        self.video_thread.start()


    def _on_video_frame(self, pixmap, result):
        """接收视频帧（优化版：减少 UI 线程阻塞）"""
        # 视频模式使用 FastTransformation 减少缩放开销
        self._fit_image(pixmap, fast=True)
        self.current_video_frame = pixmap
        self.current_video_result = result

        if result:
            self.current_result = result

            # 降低表格更新频率：每 3 帧更新一次，减少 UI 重绘
            if not hasattr(self, '_table_update_counter'):
                self._table_update_counter = 0
            self._table_update_counter += 1
            if self._table_update_counter >= 3:
                self._table_update_counter = 0
                self._update_result_table(result)

            # 语音反馈：仅在缺陷状态变化时播报（防刷屏）
            has_defect = result.has_defect
            if not hasattr(self, '_last_video_defect') or self._last_video_defect != has_defect:
                self._last_video_defect = has_defect
                self.feedback.voice.speak_result(result)
                # Web 推送改为后台线程，避免阻塞 UI
                threading.Thread(target=self._push_feedback_web, args=(result,), daemon=True).start()

        # 通知视频线程：帧已处理完毕，可以发送下一帧
        if self.video_thread:
            self.video_thread.on_frame_processed()

    def _on_fps_update(self, fps):
        """更新 FPS"""
        self.fps_label.setText(f"FPS: {fps:.1f}")

    def _on_video_error(self, error_msg):
        """视频错误"""
        QMessageBox.critical(self, "视频流错误", error_msg)
        self._stop_video()

    def _on_video_stream_finished(self):
        """视频流结束"""
        self._stop_video()
        self.status_label.setText("视频播放完毕")

    def _take_snapshot(self):
        """拍照统计 — 记录到拍照面板并自动保存标注图片"""
        if not self.is_video_streaming:
            return

        if not hasattr(self, 'current_video_result') or not self.current_video_result:
            QMessageBox.information(self, "提示", "当前帧无检测结果")
            return

        result = self.current_video_result
        self.all_results.append(result)
        self._update_stats()
        self._update_defect_distribution()

        # 添加到拍照记录表
        row = self.snapshot_table.rowCount()
        self.snapshot_table.insertRow(0)
        now = datetime.now().strftime("%H:%M:%S")
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        verdict = "合格" if not result.has_defect else "缺陷"
        defects = ", ".join(set(d.class_name_cn for d in result.detections)) or "-"

        for col, text in enumerate([now, verdict, defects]):
            item = QTableWidgetItem(text)
            item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled)
            if verdict == "缺陷":
                item.setForeground(QColor(COLOR_ERR))
            else:
                item.setForeground(QColor(COLOR_OK))
            self.snapshot_table.setItem(0, col, item)

        # 自动保存标注后的图片
        saved_path = self._save_snapshot_annotated(timestamp, verdict, defects, result)

        # 语音反馈 + Web 推送
        self.feedback.voice.speak_result(result)
        self._push_feedback_web(result)

        total = len(self.all_results)
        status = "缺陷" if result.has_defect else "合格"
        if saved_path:
            self.status_label.setText(f"拍照: {status} | 总计 {total} 张 | 已保存: {Path(saved_path).name}")
        else:
            self.status_label.setText(f"拍照: {status} | 总计 {total} 张")

    def _save_snapshot_annotated(self, timestamp, verdict, defects, result):
        """保存带标注的拍照图片"""
        try:
            # 获取当前视频帧
            if not hasattr(self, 'current_video_frame') or not self.current_video_frame:
                return None

            # 创建保存目录
            save_dir = Path(self.settings["auto_save_dir"]) / "snapshots"
            save_dir.mkdir(parents=True, exist_ok=True)

            # 生成文件名
            defect_tag = defects.replace(", ", "_").replace(" ", "") if defects != "-" else "pass"
            filename = f"{timestamp}_{verdict}_{defect_tag}.jpg"
            save_path = save_dir / filename

            # 将 QPixmap 转换为 OpenCV 格式
            pixmap = self.current_video_frame
            image = pixmap.toImage()
            width = image.width()
            height = image.height()

            # 获取图像数据
            ptr = image.bits()
            ptr.setsize(height * width * 4)
            arr = np.frombuffer(ptr, np.uint8).reshape((height, width, 4))

            # 转换 BGRA -> BGR
            frame = cv2.cvtColor(arr, cv2.COLOR_BGRA2BGR)

            # 绘制检测框和标签（带中文支持）
            annotated = frame.copy()
            for det in result.detections:
                x1, y1, x2, y2 = [int(c) for c in det.bbox]
                color = (0, 0, 255) if det.is_defect else (0, 200, 0)
                cv2.rectangle(annotated, (x1, y1), (x2, y2), color, 2)

                label = f"{det.class_name_cn} {det.confidence:.0%}"
                annotated = _put_cn_text(
                    annotated, label,
                    pos=(x1 + 2, y1 - 22),
                    font_size=16,
                    color=(255, 255, 255),
                    bg_color=color,
                )

            # 保存标注图片
            cv2.imwrite(str(save_path), annotated)

            return str(save_path)

        except Exception as e:
            print(f"[拍照保存失败] {e}")
            return None

    def _stop_video(self):
        """停止视频流"""
        if not self.is_video_streaming:
            return

        if self.video_thread:
            self.video_thread.stop()
            # 增加超时到 5 秒，给推理足够时间完成
            if not self.video_thread.wait(5000):
                # 超时强制终止
                self.video_thread.terminate()
                self.video_thread.wait(1000)
            self.video_thread = None

        self.is_video_streaming = False
        self._set_video_ui_state(False)

        self.fps_label.setText("FPS: --")
        self.image_info_label.setText("视频流已停止")
        self.status_label.setText("视频流检测已停止")

    def _set_video_ui_state(self, streaming):
        """设置视频 UI 状态 — 工具栏 + 右侧面板同步"""
        mode_idx = self.mode_combo.currentIndex()
        is_camera = (mode_idx == 1)
        is_video = (mode_idx == 2)

        # 工具栏
        self.camera_combo.setVisible(not streaming and is_camera)
        self.camera_btn.setVisible(not streaming and is_camera)
        self.phone_camera_btn.setVisible(not streaming and is_camera)
        self.video_file_btn.setVisible(not streaming and is_video)
        self.stop_video_btn.setVisible(streaming)
        self.snapshot_btn.setVisible(streaming)
        self.fps_label.setVisible(streaming)
        self.mode_combo.setEnabled(not streaming)

        # 右侧面板
        self.right_mode_combo.setEnabled(not streaming)
        self.right_camera_combo.setVisible(not streaming and is_camera)
        self.right_camera_btn.setVisible(not streaming and is_camera)
        self.right_phone_btn.setVisible(not streaming and is_camera)
        self.right_video_btn.setVisible(not streaming and is_video)
        self.right_stop_btn.setVisible(streaming)
        self.right_panel_widgets["snapshot_frame"].setVisible(streaming and is_camera)
        self.right_panel_widgets["snapshot_btn"].setVisible(streaming and is_camera)

        # 菜单
        self.camera_menu_action.setEnabled(not streaming)
        self.phone_camera_menu_action.setEnabled(not streaming)
        self.video_file_menu_action.setEnabled(not streaming)
        self.stop_video_menu_action.setEnabled(streaming)
        self.snapshot_menu_action.setEnabled(streaming)

        self.start_btn.setEnabled(not streaming and bool(self.current_image_path))
        self.start_btn2.setEnabled(not streaming and bool(self.current_image_path))

    def _toggle_timer_mode(self):
        """切换计时器模式"""
        if self.timer_mode:
            self.timer_mode = False
            self.timer_paused = False
            self.countdown_timer.stop()
            self.timer_label.hide()
            self.timer_pause_btn.hide()
            self.timer_finish_btn.hide()
            self.timer_btn.setText("  180秒计时")
            self.status_label.setText("计时模式已关闭")
        else:
            self.timer_mode = True
            self.timer_paused = False
            self.timer_remaining = self.timer_seconds
            self._update_timer_display()
            self.timer_label.show()
            self.timer_pause_btn.show()
            self.timer_pause_btn.setText("暂停")
            self.timer_finish_btn.show()
            self.countdown_timer.start(1000)

    def _toggle_timer_pause(self):
        """暂停/继续计时"""
        if not self.timer_mode:
            return

        if self.timer_paused:
            # 继续计时
            self.timer_paused = False
            self.countdown_timer.start(1000)
            self.timer_pause_btn.setText("暂停")
            self.status_label.setText("计时继续")
        else:
            # 暂停计时
            self.timer_paused = True
            self.countdown_timer.stop()
            self.timer_pause_btn.setText("继续")
            self.status_label.setText("计时已暂停")

    def _finish_timer(self):
        """完成按钮：提前结束计时并统计"""
        if not self.timer_mode:
            return
        self.countdown_timer.stop()
        
        # 停止批量检测（如果正在运行）
        if self.is_detecting and self.detection_thread:
            self.detection_thread.stop()
            self.detection_thread.wait(1000)  # 等待线程结束
            self.is_detecting = False
            self._batch_timer.stop()
        
        elapsed = self.timer_seconds - self.timer_remaining
        self._show_timer_stats(elapsed)
        self.timer_mode = False
        self.timer_paused = False
        self.timer_label.hide()
        self.timer_pause_btn.hide()
        self.timer_finish_btn.hide()
        self.timer_btn.setText("  180秒计时")

    def _show_timer_stats(self, elapsed):
        """弹出计时统计对话框"""
        try:
            total = len(self.all_results)
            pass_count = sum(1 for r in self.all_results if not r.has_defect)
            fail_count = total - pass_count

            # 统计缺陷类型
            defect_types = {}
            for r in self.all_results:
                if r.has_defect and hasattr(r, 'detections'):
                    for det in r.detections:
                        name = det.class_name if hasattr(det, 'class_name') else '未知'
                        defect_types[name] = defect_types.get(name, 0) + 1

            # 创建并显示自定义对话框
            dialog = StatsDialog(
                parent=self,
                elapsed=elapsed,
                total_seconds=self.timer_seconds,
                total=total,
                pass_count=pass_count,
                fail_count=fail_count,
                defect_types=defect_types
            )
            dialog.exec_()
            
        except Exception as e:
            QMessageBox.warning(self, "统计错误", f"统计时发生错误: {str(e)}")

    def _update_timer(self):
        """更新计时器"""
        self.timer_remaining -= 1
        self._update_timer_display()

        if self.timer_remaining <= 0:
            self.countdown_timer.stop()
            self._timer_finished()

    def _update_timer_display(self):
        """更新计时器显示"""
        self.timer_label.setText(f"{self.timer_remaining}s")

        if self.timer_remaining <= 30:
            self.timer_label.setStyleSheet(f"""
                font-size: 14px;
                font-weight: 700;
                color: {COLOR_ERR};
            """)
        else:
            self.timer_label.setStyleSheet(f"""
                font-size: 14px;
                font-weight: 600;
                color: {COLOR_WARN};
            """)

    def _timer_finished(self):
        """计时器结束"""
        self.timer_mode = False
        self.timer_label.hide()
        self.timer_pause_btn.hide()
        self.timer_finish_btn.hide()
        self._show_timer_stats(self.timer_seconds)

    def _export_report(self):
        """导出报告"""
        if not self.all_results:
            QMessageBox.information(self, "提示", "没有检测结果可导出")
            return

        file_path, file_type = QFileDialog.getSaveFileName(
            self, "导出报告", "",
            "Excel 文件 (*.xlsx);;CSV 文件 (*.csv);;文本文件 (*.txt)"
        )

        if not file_path:
            return

        try:
            if file_path.endswith('.xlsx'):
                self._export_xlsx(file_path)
            elif file_path.endswith('.csv'):
                self._export_csv(file_path)
            else:
                self._export_txt(file_path)

            self.status_label.setText(f"报告已导出: {Path(file_path).name}")
            QMessageBox.information(self, "成功", f"报告已导出到:\n{file_path}")

        except Exception as e:
            QMessageBox.critical(self, "错误", f"导出失败:\n{str(e)}")

    def _export_txt(self, file_path):
        """导出 TXT 报告"""
        with open(file_path, 'w', encoding='utf-8') as f:
            f.write("=" * 60 + "\n")
            f.write("钢材缺陷检测报告\n")
            f.write(f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write("=" * 60 + "\n\n")

            total = len(self.all_results)
            pass_count = sum(1 for r in self.all_results if not r.has_defect)
            fail_count = total - pass_count
            rate = (pass_count / total * 100) if total > 0 else 0

            f.write(f"检测总数: {total}\n")
            f.write(f"合格数量: {pass_count}\n")
            f.write(f"不合格数量: {fail_count}\n")
            f.write(f"合格率: {rate:.1f}%\n\n")

            f.write("详细结果:\n")
            f.write("-" * 60 + "\n")

            for i, result in enumerate(self.all_results, 1):
                defect_types = ', '.join(set(d.class_name_cn for d in result.detections))
                f.write(f"{i}. {Path(result.image_path).name} | "
                        f"{'合格' if not result.has_defect else '不合格'} | "
                        f"缺陷: {defect_types or '无'} | "
                        f"{result.inference_time_ms:.1f}ms\n")

    def _export_xlsx(self, file_path):
        """导出 Excel 报告"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = "检测报告"

        # 表头样式
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        # 合格/不合格颜色
        pass_font = Font(color="228B22")
        fail_font = Font(color="CC0000", bold=True)

        headers = ['序号', '图片', '判定', '缺陷数量', '缺陷类型', '置信度', '严重程度', '推理时间(ms)']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        for i, result in enumerate(self.all_results, 1):
            row = i + 1
            defect_types = ', '.join(set(d.class_name_cn for d in result.detections))
            conf_str = ', '.join(f"{d.confidence:.2%}" for d in result.detections) if result.detections else '-'
            severity, sev_score = db.compute_overall_severity(result.detections)
            verdict = '合格' if not result.has_defect else '不合格'

            values = [
                i,
                Path(result.image_path).name,
                verdict,
                len(result.detections),
                defect_types or '-',
                conf_str,
                severity,
                round(result.inference_time_ms, 1),
            ]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center" if col in (1, 3, 4, 7, 8) else "left")
                # 判定列着色
                if col == 3:
                    cell.font = pass_font if verdict == '合格' else fail_font
                # 严重程度着色
                if col == 7:
                    color_map = {'严重': 'CC0000', '中等': 'CC8800', '轻微': '228B22', '合格': '228B22'}
                    cell.font = Font(color=color_map.get(severity, '000000'), bold=(severity == '严重'))

        # 列宽自适应
        widths = [6, 30, 10, 10, 20, 16, 10, 14]
        for col, w in enumerate(widths, 1):
            ws.column_dimensions[chr(64 + col)].width = w

        # 冻结首行
        ws.freeze_panes = "A2"

        wb.save(file_path)

    def _export_csv(self, file_path):
        """导出 CSV 报告"""
        import csv
        with open(file_path, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            writer.writerow(['序号', '图片', '判定', '缺陷数量', '缺陷类型', '置信度', '严重程度', '推理时间(ms)'])

            for i, result in enumerate(self.all_results, 1):
                defect_types = ', '.join(set(d.class_name_cn for d in result.detections))
                conf_str = ', '.join(f"{d.confidence:.2%}" for d in result.detections) if result.detections else '-'
                severity, _ = db.compute_overall_severity(result.detections)
                writer.writerow([
                    i,
                    Path(result.image_path).name,
                    '合格' if not result.has_defect else '不合格',
                    len(result.detections),
                    defect_types or '-',
                    conf_str,
                    severity,
                    f"{result.inference_time_ms:.1f}"
                ])

    def _open_settings(self):
        """打开设置对话框"""
        dialog = SettingsDialog(self.settings, self)
        if dialog.exec_() == QDialog.Accepted:
            self.settings = dialog.get_settings()
            # 同步更新反馈管理器
            self.feedback.voice.enabled = self.settings["enable_voice"]
            if self.settings["enable_voice"] and self.feedback.voice._engine is None:
                self.feedback.voice._init_engine()

    def _toggle_auto_save(self, checked):
        """切换自动保存缺陷图"""
        self.settings["auto_save_defects"] = checked
        status = "开启" if checked else "关闭"
        self.status_label.setText(f"自动保存缺陷图: {status}")

    def _start_web_monitor(self):
        """启动 Web 监控服务"""
        if self.web_thread and self.web_thread.is_alive():
            QMessageBox.information(self, "提示", "Web 监控已在运行中")
            return

        def run_web():
            try:
                from web_dashboard import start_web
                start_web(port=self.settings["web_port"])
            except Exception as e:
                print(f"Web 监控启动失败: {e}")

        self.web_thread = threading.Thread(target=run_web, daemon=True)
        self.web_thread.start()
        port = self.settings["web_port"]
        self.status_label.setText(f"Web 监控已启动: http://localhost:{port}")
        QMessageBox.information(
            self, "Web 监控",
            f"Web 监控已启动！\n\n"
            f"本机访问: http://localhost:{port}\n"
            f"局域网访问: http://<本机IP>:{port}\n\n"
            f"同一网络下的设备可通过浏览器访问。"
        )

    def _open_web_browser(self):
        """在浏览器中打开 Web 监控"""
        port = self.settings["web_port"]
        webbrowser.open(f"http://localhost:{port}")

    def _open_defect_folder(self):
        """打开缺陷记录文件夹"""
        save_dir = Path(self.settings["auto_save_dir"])
        save_dir.mkdir(parents=True, exist_ok=True)
        os.startfile(str(save_dir))

    def _show_about(self):
        """显示关于对话框"""
        QMessageBox.about(
            self, "关于",
            f"<h2>{APP_TITLE_CN}</h2>"
            f"<p>版本: {APP_VERSION}</p>"
            f"<p>基于 YOLOv8 的钢材表面缺陷检测系统</p>"
            f"<p>支持 6 种缺陷类型检测</p>"
            f"<hr>"
            f"<p>功能特性:</p>"
            f"<ul>"
            f"<li>图片/文件夹批量检测</li>"
            f"<li>摄像头实时检测</li>"
            f"<li>📱 手机摄像头视频流检测</li>"
            f"<li>视频文件检测</li>"
            f"<li>✂️ 自选区缩放导出 6MP 图片</li>"
            f"<li>180秒计时演示模式</li>"
            f"<li>缺陷严重程度分级</li>"
            f"<li>自动保存缺陷图片</li>"
            f"<li>检测历史数据库</li>"
            f"<li>远程 Web 监控</li>"
            f"<li>批量导出标注图</li>"
            f"<li>报警阈值按类别可调</li>"
            f"<li>最近打开记录</li>"
            f"</ul>"
            f"<hr>"
            f"<p>📱 手机摄像头使用方法:</p>"
            f"<ol>"
            f"<li>手机安装 IP Webcam 应用</li>"
            f"<li>手机和电脑连接同一 WiFi</li>"
            f"<li>打开 IP Webcam，点击 Start server</li>"
            f"<li>在本软件中点击 📱 手机摄像头</li>"
            f"<li>输入手机显示的 IP 地址</li>"
            f"</ol>"
        )

    # ── Web 监控 ──

    def _start_web_monitor(self):
        """启动 Web 监控服务"""
        if self.web_thread and self.web_thread.is_alive():
            QMessageBox.information(self, "提示", "Web 监控已在运行中")
            self._open_web_browser()
            return

        def run_web():
            try:
                from web_dashboard import start_web
                start_web(port=self.settings.get("web_port", 5000))
            except Exception as e:
                print(f"[Web 监控] 启动失败: {e}")

        self.web_thread = threading.Thread(target=run_web, daemon=True)
        self.web_thread.start()

        port = self.settings.get("web_port", 5000)
        self.status_label.setText(f"Web 监控已启动: http://localhost:{port}")

        # 延迟打开浏览器
        QTimer.singleShot(1500, self._open_web_browser)

    def _open_web_browser(self):
        """在浏览器中打开 Web 监控"""
        import webbrowser
        port = self.settings.get("web_port", 5000)
        webbrowser.open(f"http://localhost:{port}")

    def _open_defect_folder(self):
        """打开缺陷记录文件夹"""
        save_dir = Path(self.settings["auto_save_dir"])
        if save_dir.exists():
            os.startfile(str(save_dir))
        else:
            QMessageBox.information(self, "提示", "缺陷记录文件夹不存在")

    def _toggle_auto_save(self, checked):
        """切换自动保存"""
        self.settings["auto_save_defects"] = checked

    def _batch_export_annotated(self):
        """批量导出带标注的图片"""
        if not self.all_results:
            QMessageBox.information(self, "提示", "没有检测结果可导出")
            return

        dir_path = QFileDialog.getExistingDirectory(self, "选择导出目录")
        if not dir_path:
            return

        try:
            export_dir = Path(dir_path) / f"annotated_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            export_dir.mkdir(parents=True, exist_ok=True)

            for i, result in enumerate(self.all_results):
                try:
                    annotated = self.detector.visualize_result(result.image_path, result)
                    filename = Path(result.image_path).name
                    save_path = export_dir / f"result_{i+1:04d}_{filename}"
                    cv2.imwrite(str(save_path), annotated)
                except Exception as e:
                    print(f"导出失败 {result.image_path}: {e}")

            self.status_label.setText(f"已导出 {len(self.all_results)} 张标注图到: {export_dir}")
            QMessageBox.information(self, "成功",
                                    f"已导出 {len(self.all_results)} 张标注图片到:\n{export_dir}")

        except Exception as e:
            QMessageBox.critical(self, "错误", f"导出失败:\n{str(e)}")

    def dragEnterEvent(self, event: QDragEnterEvent):
        """拖拽进入事件"""
        if event.mimeData().hasUrls():
            event.acceptProposedAction()

    def dropEvent(self, event: QDropEvent):
        """拖拽放下事件"""
        urls = event.mimeData().urls()
        if urls:
            file_path = urls[0].toLocalFile()
            if Path(file_path).is_file():
                self.current_image_path = file_path
                self._display_image(file_path)
                self.start_btn.setEnabled(True)
                self.start_btn2.setEnabled(True)
                self.image_info_label.setText(f"文件: {Path(file_path).name}")
            elif Path(file_path).is_dir():
                self.current_image_path = file_path
                self._open_directory()


# ============================================================
# 设置对话框
# ============================================================

class SettingsDialog(QDialog):
    """设置对话框"""

    def __init__(self, settings, parent=None):
        super().__init__(parent)
        self.settings = settings.copy()
        self.setWindowTitle("检测参数设置")
        self.setMinimumWidth(400)

        layout = QFormLayout(self)

        # 置信度阈值
        self.conf_spin = QDoubleSpinBox()
        self.conf_spin.setRange(0.01, 1.0)
        self.conf_spin.setSingleStep(0.05)
        self.conf_spin.setValue(settings["conf_threshold"])
        layout.addRow("置信度阈值:", self.conf_spin)

        # IoU 阈值
        self.iou_spin = QDoubleSpinBox()
        self.iou_spin.setRange(0.01, 1.0)
        self.iou_spin.setSingleStep(0.05)
        self.iou_spin.setValue(settings["iou_threshold"])
        layout.addRow("IoU 阈值:", self.iou_spin)

        # 图片尺寸
        self.img_size_spin = QSpinBox()
        self.img_size_spin.setRange(320, 1280)
        self.img_size_spin.setSingleStep(32)
        self.img_size_spin.setValue(settings["img_size"])
        layout.addRow("图片尺寸:", self.img_size_spin)

        # 语音播报
        self.voice_check = QCheckBox("启用语音播报")
        self.voice_check.setChecked(settings.get("enable_voice", False))
        layout.addRow("语音反馈:", self.voice_check)

        # 日志记录
        self.log_check = QCheckBox("启用检测日志")
        self.log_check.setChecked(settings.get("enable_log", True))
        layout.addRow("日志记录:", self.log_check)

        # 按钮
        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addRow(buttons)

    def get_settings(self):
        """获取设置"""
        self.settings["conf_threshold"] = self.conf_spin.value()
        self.settings["iou_threshold"] = self.iou_spin.value()
        self.settings["img_size"] = self.img_size_spin.value()
        self.settings["enable_voice"] = self.voice_check.isChecked()
        self.settings["enable_log"] = self.log_check.isChecked()
        return self.settings


class ReviewDialog(QDialog):
    """低置信度人工复核对话框"""

    def __init__(self, image_path, detections, parent=None):
        super().__init__(parent)
        self.setWindowTitle("人工复核 — 低置信度检测")
        self.setMinimumWidth(450)
        self.result_data = None

        layout = QVBoxLayout(self)

        # 提示
        info = QLabel("以下检测结果置信度较低，请确认或修正：")
        info.setStyleSheet("color: #fbbf24; font-weight: bold; font-size: 13px;")
        layout.addWidget(info)

        # 检测结果列表
        self.detections = detections
        self.combos = []
        for det in detections:
            row = QHBoxLayout()
            label = QLabel(f"{det.class_name_cn} ({det.confidence:.0%})")
            label.setStyleSheet("color: #e0e0e0; font-size: 12px;")
            row.addWidget(label)

            combo = QComboBox()
            from constants import CLASS_NAMES_CN
            class_names = list(CLASS_NAMES_CN.keys())
            cn_names = list(CLASS_NAMES_CN.values())
            combo.addItems(cn_names)
            # 默认选中当前检测类别
            if det.class_name_cn in cn_names:
                combo.setCurrentIndex(cn_names.index(det.class_name_cn))
            combo.setStyleSheet("background: #222; color: #e0e0e0; padding: 4px; border-radius: 4px;")
            row.addWidget(combo)
            self.combos.append(combo)
            layout.addLayout(row)

        # 按钮
        btn_layout = QHBoxLayout()

        confirm_btn = QPushButton("全部确认")
        confirm_btn.setStyleSheet("background: #22c55e; color: white; padding: 8px 16px; border-radius: 4px; font-weight: bold;")
        confirm_btn.clicked.connect(self._confirm_all)
        btn_layout.addWidget(confirm_btn)

        correct_btn = QPushButton("应用修正")
        correct_btn.setStyleSheet("background: #4a9eff; color: white; padding: 8px 16px; border-radius: 4px; font-weight: bold;")
        correct_btn.clicked.connect(self._apply_correction)
        btn_layout.addWidget(correct_btn)

        skip_btn = QPushButton("跳过")
        skip_btn.setStyleSheet("background: #555; color: #ccc; padding: 8px 16px; border-radius: 4px;")
        skip_btn.clicked.connect(self.reject)
        btn_layout.addWidget(skip_btn)

        layout.addLayout(btn_layout)

    def _confirm_all(self):
        """确认所有检测结果"""
        self.result_data = [(d, d.class_name, d.class_name_cn, "确认") for d in self.detections]
        self.accept()

    def _apply_correction(self):
        """应用修正"""
        from constants import CLASS_NAMES_CN
        cn_to_en = {v: k for k, v in CLASS_NAMES_CN.items()}
        self.result_data = []
        for i, det in enumerate(self.detections):
            selected_cn = self.combos[i].currentText()
            selected_en = cn_to_en.get(selected_cn, det.class_name)
            action = "确认" if selected_cn == det.class_name_cn else "修正"
            self.result_data.append((det, selected_en, selected_cn, action))
        self.accept()


class PhoneStreamDialog(QDialog):
    """手机视频流连接对话框"""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("连接手机摄像头")
        self.setMinimumWidth(500)

        layout = QVBoxLayout(self)

        # 说明文字
        info_label = QLabel(
            "📱 使用手机摄像头进行视频流检测\n\n"
            "步骤：\n"
            "1. 在手机上安装 IP Webcam 应用（Google Play 可下载）\n"
            "2. 打开应用，点击底部 'Start server'\n"
            "3. 应用会显示一个 IP 地址（如 192.168.1.100:8080）\n"
            "4. 将地址填入下方输入框\n\n"
            "支持格式：\n"
            "• RTSP: rtsp://192.168.1.100:8080/h264_pcm.sdp\n"
            "• HTTP: http://192.168.1.100:8080/video\n"
            "• MJPEG: http://192.168.1.100:8080/videofeed"
        )
        info_label.setStyleSheet(f"color: {COLOR_TEXT}; padding: 10px;")
        info_label.setWordWrap(True)
        layout.addWidget(info_label)

        # 地址输入
        form_layout = QFormLayout()
        self.url_input = QComboBox()
        self.url_input.setEditable(True)
        self.url_input.setPlaceholderText("输入手机视频流地址...")
        self.url_input.addItems([
            "http://192.168.1.100:8080/video",
            "http://192.168.1.100:8080/videofeed",
            "rtsp://192.168.1.100:8080/h264_pcm.sdp",
        ])
        form_layout.addRow("视频流地址:", self.url_input)
        layout.addLayout(form_layout)

        # 分辨率选择
        res_layout = QFormLayout()
        self.resolution_combo = QComboBox()
        self.resolution_combo.addItems([
            "原始分辨率",
            "限制为 6MP (3072×2048)",
            "限制为 1080p (1920×1080)",
            "限制为 720p (1280×720)",
        ])
        res_layout.addRow("输出分辨率:", self.resolution_combo)
        layout.addLayout(res_layout)

        # 按钮
        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

    def get_url(self):
        """获取输入的 URL"""
        return self.url_input.currentText().strip()

    def get_resolution_limit(self):
        """获取分辨率限制"""
        text = self.resolution_combo.currentText()
        if "6MP" in text:
            return (3072, 2048)
        elif "1080p" in text:
            return (1920, 1080)
        elif "720p" in text:
            return (1280, 720)
        return None  # 原始分辨率


# ============================================================
# 工作线程
# ============================================================

class VideoStreamThread(QThread):
    """视频流检测线程（优化版：帧率控制 + 批量绘制 + 资源保护）"""

    frame_ready = pyqtSignal(object, object)
    fps_update = pyqtSignal(float)
    error_occurred = pyqtSignal(str)
    stream_finished = pyqtSignal()

    # 目标帧率限制，防止信号风暴淹没 UI 线程
    TARGET_FPS = 30
    FRAME_INTERVAL = 1.0 / TARGET_FPS

    def __init__(self, detector, source=0, parent=None, resolution_limit=None, grayscale_enabled=False):
        super().__init__(parent)
        self.detector = detector
        self.source = source
        self.resolution_limit = resolution_limit  # (width, height) 或 None
        self.grayscale_enabled = grayscale_enabled
        self._is_running = True
        self._frame_pending = False  # 跳帧标志：UI 还未处理完上一帧

    def _limit_resolution(self, frame):
        """限制帧分辨率"""
        if self.resolution_limit is None:
            return frame

        target_w, target_h = self.resolution_limit
        h, w = frame.shape[:2]

        # 如果已经在目标分辨率内，不处理
        if w <= target_w and h <= target_h:
            return frame

        # 计算缩放比例
        scale = min(target_w / w, target_h / h)
        new_w = int(w * scale)
        new_h = int(h * scale)

        return cv2.resize(frame, (new_w, new_h), interpolation=cv2.INTER_LINEAR)

    def run(self):
        cap = None
        try:
            # 手机视频流通常需要特殊处理
            if isinstance(self.source, str) and (
                self.source.startswith('rtsp://') or
                self.source.startswith('http://') or
                self.source.startswith('https://')
            ):
                # 网络流：设置缓冲区大小减少延迟
                cap = cv2.VideoCapture(self.source, cv2.CAP_FFMPEG)
                cap.set(cv2.CAP_PROP_BUFFERSIZE, 1)
                retry_count = 0
                max_retries = 30  # 最大重试次数
            elif isinstance(self.source, int):
                cap = cv2.VideoCapture(self.source, cv2.CAP_DSHOW)
                if not cap.isOpened():
                    cap.release()
                    cap = cv2.VideoCapture(self.source, cv2.CAP_ANY)
                retry_count = 0
                max_retries = 0  # 摄像头不重试
            else:
                cap = cv2.VideoCapture(self.source)
                retry_count = 0
                max_retries = 0

            if not cap.isOpened():
                self.error_occurred.emit(f"无法打开视频源: {self.source}\n\n请检查：\n1. 手机和电脑是否在同一网络\n2. IP Webcam 是否已启动\n3. 地址是否正确")
                return

            fps_counter = 0
            fps_start_time = time.perf_counter()

            while self._is_running and cap.isOpened():
                frame_start = time.perf_counter()

                ret, frame = cap.read()
                if not ret:
                    if isinstance(self.source, str):
                        # 网络流重试逻辑
                        retry_count += 1
                        if max_retries > 0 and retry_count > max_retries:
                            self.error_occurred.emit(f"视频流重试 {max_retries} 次后失败，请检查网络连接")
                            break
                        time.sleep(0.1)
                        continue
                    break
                else:
                    retry_count = 0  # 读取成功，重置重试计数

                # 跳帧机制：如果 UI 还没处理完上一帧，跳过当前帧
                if self._frame_pending:
                    continue

                # 限制分辨率
                frame = self._limit_resolution(frame)

                # 灰度转换
                if self.grayscale_enabled:
                    gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
                    frame = cv2.cvtColor(gray, cv2.COLOR_GRAY2BGR)

                try:
                    t_start = time.perf_counter()
                    results = self.detector.model.predict(
                        frame,
                        conf=self.detector.conf_threshold,
                        iou=self.detector.iou_threshold,
                        imgsz=self.detector.img_size,
                        device=self.detector.device,
                        half=self.detector.use_fp16,
                        verbose=False,
                        max_det=50,  # 限制最大检测数，减少 NMS 时间（钢材缺陷通常不超过 20 个）
                    )
                    t_end = time.perf_counter()
                    inference_time_ms = (t_end - t_start) * 1000

                    result = self.detector._parse_results(results[0], "<video_frame>", frame.shape)
                    result.inference_time_ms = inference_time_ms

                    # 批量绘制检测框（性能优化：只做一次颜色空间转换）
                    if result.detections:
                        annotated = draw_detections_cn(frame, result.detections, font_size=16)
                    else:
                        annotated = frame

                    # 转换为 QPixmap
                    rgb_img = cv2.cvtColor(annotated, cv2.COLOR_BGR2RGB)
                    h, w, ch = rgb_img.shape
                    bytes_per_line = ch * w
                    q_img = QImage(rgb_img.data, w, h, bytes_per_line, QImage.Format_RGB888).copy()
                    pixmap = QPixmap.fromImage(q_img)

                    self._frame_pending = True  # 标记帧等待处理
                    self.frame_ready.emit(pixmap, result)

                except Exception as e:
                    rgb_img = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
                    h, w, ch = rgb_img.shape
                    bytes_per_line = ch * w
                    q_img = QImage(rgb_img.data, w, h, bytes_per_line, QImage.Format_RGB888).copy()
                    pixmap = QPixmap.fromImage(q_img)
                    self._frame_pending = True
                    self.frame_ready.emit(pixmap, None)

                fps_counter += 1
                elapsed = time.perf_counter() - fps_start_time
                if elapsed >= 0.5:
                    current_fps = fps_counter / elapsed
                    self.fps_update.emit(current_fps)
                    fps_counter = 0
                    fps_start_time = time.perf_counter()

                # 帧率限制：确保每帧至少间隔 FRAME_INTERVAL
                frame_elapsed = time.perf_counter() - frame_start
                if frame_elapsed < self.FRAME_INTERVAL:
                    time.sleep(self.FRAME_INTERVAL - frame_elapsed)

        except Exception as e:
            self.error_occurred.emit(f"视频流异常: {e}")
        finally:
            # 确保资源释放
            if cap is not None:
                cap.release()
            self.stream_finished.emit()

    def on_frame_processed(self):
        """UI 线程调用，通知帧已处理完毕"""
        self._frame_pending = False

    def stop(self):
        self._is_running = False


class DetectionThread(QThread):
    """批量检测线程"""

    progress = pyqtSignal(int, int)
    result_ready = pyqtSignal(object, object)
    finished_all = pyqtSignal(object)
    error_occurred = pyqtSignal(str)

    def __init__(self, detector, image_paths, parent=None, grayscale_enabled=False):
        super().__init__(parent)
        self.detector = detector
        self.image_paths = image_paths
        self.grayscale_enabled = grayscale_enabled
        self._is_running = True

    def run(self):
        last_result = None
        for i, image_path in enumerate(self.image_paths):
            if not self._is_running:
                break

            try:
                if self.grayscale_enabled:
                    img = cv2.imread(image_path)
                    if img is not None:
                        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
                        img = cv2.cvtColor(gray, cv2.COLOR_GRAY2BGR)
                        result = self.detector.detect_single(img)
                    else:
                        result = self.detector.detect_single(image_path)
                else:
                    result = self.detector.detect_single(image_path)
                last_result = result
                self.result_ready.emit(image_path, result)
                self.progress.emit(i + 1, len(self.image_paths))
            except Exception as e:
                self.error_occurred.emit(f"检测失败 {image_path}: {str(e)}")

        self.finished_all.emit(last_result)

    def stop(self):
        self._is_running = False


# ============================================================
# 启动入口
# ============================================================

def main():
    app = QApplication(sys.argv)

    # 设置全局字体
    font = QFont("Microsoft YaHei", 10)
    app.setFont(font)

    # 应用样式
    app.setStyleSheet(MODERN_STYLE)

    window = SteelDefectApp()
    window.show()

    sys.exit(app.exec_())


if __name__ == "__main__":
    main()
